import io
from datetime import date, timedelta
from decimal import Decimal
from uuid import UUID

from django.contrib.auth import get_user_model
from django.core.cache import cache
from django.db import transaction
from django.db.models import Count, DecimalField, F, Max, Q, Sum
from django.http import FileResponse, Http404, HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from drf_spectacular.utils import OpenApiExample, OpenApiParameter, extend_schema, extend_schema_view
from rest_framework import generics, permissions
from rest_framework.exceptions import ValidationError
from rest_framework.response import Response

from apps.dashboard.backups import create_backup, delete_backup, get_auto_config, set_auto_config

from apps.accounts.filters import filter_masters_by_specialization, specialization_counts
from apps.accounts.models import Client, Master, MasterApprovalStatus
from apps.accounts.permissions import IsStaffOrAdminUser
from apps.common.filters import filter_by_category
from apps.common.responses import success_response
from apps.common.views import EnvelopeMixin
from apps.dashboard.models import (
    DashboardBackup,
    DashboardCompanyExpense,
    DashboardIntegrationSetting,
    DashboardLiveStream,
    DashboardOrderAssistant,
    DashboardWarehouseExpense,
)
from apps.market.models import MarketCategory, MarketOrder, MarketProduct, MarketProductImage
from apps.notifications.models import Notification
from apps.notifications.services import broadcast_notification, send_push_notification
from apps.orders.models import Order, OrderStatus
from apps.profiles.models import Tariff, TariffFeature
from apps.services.models import Service, ServiceCategory, ServicePrice
from apps.support.models import SupportChat, SupportMessage
from apps.support.services import (
    attach_latest_support_messages,
    broadcast_support_message,
    get_or_create_support_chat,
    mark_support_thread_read_by_admin,
    touch_chat,
    with_latest_support_message,
)
from apps.wallet.models import MasterExpense, MasterWallet, WalletTransaction, WithdrawRequest
from apps.wallet.services import (
    accept_cash_handover,
    post_wallet_transaction,
    reject_cash_handover,
    reverse_wallet_transaction,
)
from apps.warehouse.models import MasterInventory, StockMovement, WarehouseCategory, WarehouseProduct
from apps.warehouse.services import (
    adjust_master_inventory,
    adjust_warehouse_stock,
    apply_stock_movement_effect,
    return_inventory_to_warehouse,
    reverse_stock_movement,
)
from .serializers import (
    DashboardCashHandoverActionSerializer,
    DashboardCashHandoverEnvelopeSerializer,
    DashboardCashHandoverSerializer,
    DashboardCompanyExpenseSerializer,
    DashboardClientMiniSerializer,
    DashboardClientSerializer,
    DashboardExpenseSerializer,
    DashboardBackupSerializer,
    DashboardBackupSettingsSerializer,
    DashboardIntegrationSettingSerializer,
    DashboardLiveStreamSerializer,
    DashboardLoginSerializer,
    DashboardLogoutSerializer,
    DashboardMarketCategorySerializer,
    DashboardMarketOrderSerializer,
    DashboardMarketProductImageSerializer,
    DashboardMarketProductSerializer,
    DashboardMasterMiniSerializer,
    DashboardMasterLocationSerializer,
    DashboardMasterInventorySerializer,
    DashboardMasterInventoryAssignSerializer,
    DashboardMasterApplicationSerializer,
    DashboardMasterBlockSerializer,
    DashboardMasterSerializer,
    DashboardMasterStatusSerializer,
    DashboardMasterWalletSerializer,
    DashboardMeSerializer,
    DashboardNotificationSerializer,
    DashboardOrderAssistantSerializer,
    DashboardOrderAssignSerializer,
    DashboardOrderSerializer,
    DashboardOrderStatusSerializer,
    DashboardQuerySerializer,
    DashboardRefreshSerializer,
    DashboardServicePriceSerializer,
    DashboardServiceCategorySerializer,
    DashboardServiceSerializer,
    DashboardStaffSerializer,
    DashboardStockMovementSerializer,
    DashboardSupportMessageSerializer,
    DashboardTariffFeatureSerializer,
    DashboardTariffSerializer,
    DashboardWarehouseCategorySerializer,
    DashboardWarehouseExpenseSerializer,
    DashboardWarehouseProductSerializer,
    DashboardWalletTransactionSerializer,
    DashboardWithdrawRequestSerializer,
    DateRangeQuerySerializer,
    EmptySerializer,
)
from .services import (
    WEEKDAY_UZ,
    default_target_date,
    format_uzs,
    get_finance_summary,
    get_income_by_service,
    get_income_dynamics,
    get_income_expense,
    get_orders_by_service,
    get_stats,
    get_top_clients,
    get_weekly_orders,
    money,
)


DASHBOARD_AUTH_TAG = "Dashboard - Auth"
DASHBOARD_OVERVIEW_TAG = "Dashboard - Dashboard"
DASHBOARD_CLIENTS_TAG = "Dashboard - Mijozlar"
DASHBOARD_STAFF_TAG = "Dashboard - Xodim"
DASHBOARD_MAPS_TAG = "Dashboard - Xaritalar"
DASHBOARD_MASTERS_TAG = "Dashboard - Ustalar"
DASHBOARD_ORDERS_TAG = "Dashboard - Buyurtmalar"
DASHBOARD_SERVICES_TAG = "Dashboard - Xizmatlar va Narxlar"
DASHBOARD_TARIFFS_TAG = "Dashboard - Tariflar"
DASHBOARD_NOTIFICATIONS_TAG = "Dashboard - Bildirishnomalar"
DASHBOARD_EXPENSES_TAG = "Dashboard - Xarajatlar"
DASHBOARD_MARKET_TAG = "Dashboard - Marketplace"
DASHBOARD_WAREHOUSE_TAG = "Dashboard - Ombor"
DASHBOARD_LIVE_TAG = "Dashboard - Jonli Kuzatuv"
DASHBOARD_SUPPORT_TAG = "Dashboard - Xabarlar"
DASHBOARD_FINANCE_TAG = "Dashboard - Moliya Hisobotlar"
DASHBOARD_SETTINGS_TAG = "Dashboard - Sozlamalar"
ACTIVE_ORDER_STATUSES = [OrderStatus.NEW, OrderStatus.ACCEPTED, OrderStatus.ON_WAY, OrderStatus.ARRIVED]


def bool_param(value):
    if value is None:
        return None
    return str(value).lower() in {"1", "true", "yes", "on"}


def parse_uuid(value):
    try:
        return UUID(str(value))
    except (TypeError, ValueError):
        return None


ORDER_TAB_ALIASES = {
    "all": "all", "barchasi": "all",
    "yangi": "yangi", "new": "yangi",
    "yo'lda": "yo'lda", "yolda": "yo'lda", "on_way": "yo'lda",
    "bajarilmoqda": "bajarilmoqda", "in_progress": "bajarilmoqda",
    "yakunlangan": "yakunlangan", "completed": "yakunlangan",
    "bekor": "bekor", "bekor qilingan": "bekor", "cancelled": "bekor",
}


def apply_order_tab(queryset, tab):
    """Filter orders by the Figma status tab (Yangi/Yo'lda/Bajarilmoqda/Yakunlangan/Bekor)."""
    tab = ORDER_TAB_ALIASES.get((tab or "").strip().lower())
    if not tab or tab == "all":
        return queryset
    if tab == "yangi":
        # Only unassigned new orders (no active master assigned yet).
        return queryset.filter(status=OrderStatus.NEW).exclude(assigned_masters__is_active=True)
    if tab == "yo'lda":
        return queryset.filter(status=OrderStatus.ON_WAY)
    if tab == "bajarilmoqda":
        return queryset.filter(
            Q(status__in=[OrderStatus.ACCEPTED, OrderStatus.ARRIVED])
            | Q(status=OrderStatus.NEW, assigned_masters__is_active=True)
        ).distinct()
    if tab == "yakunlangan":
        return queryset.filter(status=OrderStatus.COMPLETED)
    if tab == "bekor":
        return queryset.filter(status__in=[OrderStatus.CANCELLED, OrderStatus.REJECTED])
    return queryset


class DashboardPermissionMixin:
    permission_classes = [IsStaffOrAdminUser]


@extend_schema(
    tags=[DASHBOARD_AUTH_TAG],
    summary="Dashboard login",
    description="Dashboardga kirish uchun Django staff/admin user login va parolini qabul qiladi, access_token va refresh_token qaytaradi.",
    request=DashboardLoginSerializer,
    examples=[
        OpenApiExample(
            "Login request",
            value={"username": "admin", "password": "admin-password"},
            request_only=True,
        )
    ],
)
class DashboardLoginAPIView(generics.GenericAPIView):
    # Public token endpoint: skip auth so a stray admin session cookie doesn't
    # trigger SessionAuthentication's CSRF check.
    authentication_classes = []
    permission_classes = [permissions.AllowAny]
    serializer_class = DashboardLoginSerializer

    def post(self, request):
        serializer = self.get_serializer(data=request.data, context={"request": request})
        serializer.is_valid(raise_exception=True)
        return success_response(serializer.save())


@extend_schema(
    tags=[DASHBOARD_AUTH_TAG],
    summary="Dashboard token refresh",
    description="Dashboard access token muddati tugaganda refresh_token orqali yangi access_token va refresh_token olish uchun ishlatiladi.",
    request=DashboardRefreshSerializer,
)
class DashboardRefreshAPIView(generics.GenericAPIView):
    authentication_classes = []
    permission_classes = [permissions.AllowAny]
    serializer_class = DashboardRefreshSerializer

    def post(self, request):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        return success_response(serializer.save())


@extend_schema(
    tags=[DASHBOARD_AUTH_TAG],
    summary="Dashboard logout",
    description="Dashboarddan chiqishda refresh tokenni blacklist qiladi, frontend local storage/sessiondagi tokenlarni ham tozalashi kerak.",
    request=DashboardLogoutSerializer,
)
class DashboardLogoutAPIView(DashboardPermissionMixin, generics.GenericAPIView):
    serializer_class = DashboardLogoutSerializer

    def post(self, request):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return success_response(message="Logged out")


@extend_schema(
    tags=[DASHBOARD_AUTH_TAG],
    summary="Dashboard admin profile",
    description="Hozir login bo'lgan dashboard admin/staff user ma'lumotlarini qaytaradi.",
)
class DashboardMeAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.RetrieveAPIView):
    serializer_class = DashboardMeSerializer

    def get_object(self):
        return self.request.user


@extend_schema(
    tags=[DASHBOARD_OVERVIEW_TAG],
    summary="Dashboard meta",
    description="Dashboard header va umumiy UI uchun bugungi sana, hafta kuni, timezone, til variantlari va order statuslarini qaytaradi.",
)
class DashboardMetaAPIView(DashboardPermissionMixin, generics.GenericAPIView):
    serializer_class = EmptySerializer

    def get(self, request):
        today = timezone.localdate()
        return success_response(
            {
                "date": today.isoformat(),
                "weekday": WEEKDAY_UZ[today.weekday()],
                "timezone": str(timezone.get_current_timezone()),
                "currency": "UZS",
                "languages": [
                    {"code": "ru", "label": "Russkiy", "is_active": False},
                    {"code": "uz", "label": "O'zbekcha", "is_active": True},
                    {"code": "en", "label": "English", "is_active": False},
                ],
                "order_statuses": [{"value": value, "label": label} for value, label in OrderStatus.choices],
            }
        )


@extend_schema(
    tags=[DASHBOARD_OVERVIEW_TAG],
    summary="Dashboard stats cards",
    description="Dashboard bosh sahifasidagi cardlar uchun bugungi buyurtmalar, faol ustalar, kunlik daromad va xarajat statistikalarini qaytaradi.",
    parameters=[OpenApiParameter("date", str, OpenApiParameter.QUERY, required=False)],
)
class DashboardStatsAPIView(DashboardPermissionMixin, generics.GenericAPIView):
    serializer_class = DashboardQuerySerializer

    def get(self, request):
        serializer = self.get_serializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        target_date = serializer.validated_data.get("date") or default_target_date()
        cache_key = f"dashboard:stats:{request.user.id}:{target_date}"
        data = cache.get(cache_key)
        if data is None:
            data = get_stats(target_date)
            cache.set(cache_key, data, timeout=300)
        return success_response(data)


@extend_schema(
    tags=[DASHBOARD_OVERVIEW_TAG],
    summary="Orders by service chart",
    description="Dashboard donut charti uchun xizmatlar kesimida buyurtmalar soni va foiz taqsimotini qaytaradi.",
)
class DashboardOrdersByServiceAPIView(DashboardPermissionMixin, generics.GenericAPIView):
    serializer_class = DateRangeQuerySerializer

    def get(self, request):
        serializer = self.get_serializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        date_to = serializer.validated_data.get("date_to") or default_target_date()
        date_from = serializer.validated_data.get("date_from") or date_to - timedelta(days=6)
        data = get_orders_by_service(date_from, date_to, serializer.validated_data["limit"])
        return success_response(data)


@extend_schema(
    tags=[DASHBOARD_OVERVIEW_TAG],
    summary="Weekly orders chart",
    description="Dashboarddagi 7 kunlik buyurtmalar grafigi uchun kunma-kun statuslar kesimidagi statistikani qaytaradi.",
)
class DashboardWeeklyOrdersAPIView(DashboardPermissionMixin, generics.GenericAPIView):
    serializer_class = DashboardQuerySerializer

    def get(self, request):
        serializer = self.get_serializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        target_date = serializer.validated_data.get("date") or default_target_date()
        return success_response(get_weekly_orders(target_date))


@extend_schema(
    tags=[DASHBOARD_OVERVIEW_TAG],
    summary="Income dynamics chart",
    description="Dashboard daromad dinamikasi grafigi uchun oxirgi 7 kunlik completed order daromadlarini qaytaradi.",
)
class DashboardIncomeDynamicsAPIView(DashboardPermissionMixin, generics.GenericAPIView):
    serializer_class = DashboardQuerySerializer

    def get(self, request):
        serializer = self.get_serializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        target_date = serializer.validated_data.get("date") or default_target_date()
        return success_response(get_income_dynamics(target_date))


@extend_schema(
    tags=[DASHBOARD_OVERVIEW_TAG],
    summary="Income and expense chart",
    description="Dashboard moliya grafigi uchun yil bo'yicha oyma-oy daromad va xarajat summalarini qaytaradi.",
)
class DashboardIncomeExpenseAPIView(DashboardPermissionMixin, generics.GenericAPIView):
    serializer_class = EmptySerializer

    def get(self, request):
        try:
            year = int(request.query_params.get("year", timezone.localdate().year))
        except ValueError as exc:
            raise ValidationError({"year": "year integer bo'lishi kerak"}) from exc
        return success_response(get_income_expense(year))


@extend_schema(
    tags=[DASHBOARD_OVERVIEW_TAG],
    summary="Dashboard today orders",
    description="Dashboard bosh sahifasidagi 'Bugungi faol buyurtmalar' jadvali uchun active orderlar ro'yxatini qaytaradi.",
)
class DashboardTodayOrdersAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.ListAPIView):
    serializer_class = DashboardOrderSerializer

    def get_queryset(self):
        if getattr(self, "swagger_fake_view", False):
            return Order.objects.none()
        target_date = default_target_date()
        return (
            Order.objects.select_related("client", "master", "service", "service__category", "address", "tracking")
            .filter(created_at__date=target_date, status__in=ACTIVE_ORDER_STATUSES)
            .order_by("-created_at")
        )


@extend_schema(
    tags=[DASHBOARD_OVERVIEW_TAG],
    summary="Dashboard overview",
    description="Bosh sahifa uchun barcha asosiy data: meta, admin user, stats cardlar, chartlar, bugungi orderlar va notification countni bitta response'da qaytaradi.",
)
class DashboardOverviewAPIView(DashboardPermissionMixin, generics.GenericAPIView):
    serializer_class = DashboardQuerySerializer

    def get(self, request):
        serializer = self.get_serializer(data=request.query_params)
        serializer.is_valid(raise_exception=True)
        target_date = serializer.validated_data.get("date") or default_target_date()
        orders_limit = serializer.validated_data["orders_limit"]
        date_from = target_date - timedelta(days=6)
        today_orders = (
            Order.objects.select_related("client", "master", "service", "service__category", "address", "tracking")
            .filter(created_at__date=target_date, status__in=ACTIVE_ORDER_STATUSES)
            .order_by("-created_at")[:orders_limit]
        )
        data = {
            "meta": {
                "date": target_date.isoformat(),
                "weekday": WEEKDAY_UZ[target_date.weekday()],
                "timezone": str(timezone.get_current_timezone()),
                "currency": "UZS",
            },
            "user": DashboardMeSerializer(request.user).data,
            "stats": get_stats(target_date),
            "orders_by_service": get_orders_by_service(date_from, target_date)["items"],
            "weekly_orders": get_weekly_orders(target_date)["items"],
            "income_dynamics": get_income_dynamics(target_date)["items"],
            "income_expense": get_income_expense(target_date.year)["items"][:7],
            "today_orders": {
                "count": len(today_orders),
                "results": DashboardOrderSerializer(today_orders, many=True, context={"request": request}).data,
            },
            "notifications": {
                "unread_count": Notification.objects.filter(is_read=False).count(),
            },
        }
        return success_response(data)


@extend_schema(
    tags=[DASHBOARD_OVERVIEW_TAG],
    summary="Dashboard global search",
    description="Headerdagi global qidiruv uchun xizmat, buyurtma, mijoz va usta natijalarini umumiy ro'yxat ko'rinishida qaytaradi.",
)
class DashboardSearchAPIView(DashboardPermissionMixin, generics.GenericAPIView):
    serializer_class = EmptySerializer

    def get(self, request):
        query = request.query_params.get("q", "").strip()
        if len(query) < 2:
            return success_response({"query": query, "results": []})

        results = []
        for service in Service.objects.filter(Q(name__icontains=query) | Q(description__icontains=query), is_active=True)[:5]:
            results.append(
                {"type": "service", "id": str(service.id), "title": service.name, "subtitle": "Xizmat", "url": f"/services/{service.id}"}
            )
        order_query = Q(address_text__icontains=query) | Q(note__icontains=query) | Q(client__phone__icontains=query)
        query_uuid = parse_uuid(query)
        if query_uuid:
            order_query |= Q(id=query_uuid)
        for order in Order.objects.filter(order_query).select_related("client", "service")[:5]:
            results.append(
                {
                    "type": "order",
                    "id": str(order.id),
                    "title": f"ORD-{str(order.id).split('-')[0].upper()}",
                    "subtitle": f"{order.client} - {order.service.name}",
                    "url": f"/orders/{order.id}",
                }
            )
        client_query = Q(first_name__icontains=query) | Q(last_name__icontains=query) | Q(phone__icontains=query)
        for client in Client.objects.filter(client_query)[:5]:
            results.append(
                {"type": "client", "id": str(client.id), "title": str(client), "subtitle": client.phone, "url": f"/clients/{client.id}"}
            )
        master_query = Q(first_name__icontains=query) | Q(last_name__icontains=query) | Q(phone__icontains=query)
        for master in Master.objects.filter(master_query)[:5]:
            results.append(
                {"type": "master", "id": str(master.id), "title": master.full_name, "subtitle": master.phone, "url": f"/masters/{master.id}"}
            )

        return success_response({"query": query, "results": results[:10]})


@extend_schema(
    tags=[DASHBOARD_CLIENTS_TAG],
    summary="Dashboard clients list/create",
    description="Mijozlar sahifasi uchun clientlar ro'yxatini filter/search bilan qaytaradi yoki dashboarddan yangi client yaratadi.",
)
class DashboardClientListCreateAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.ListCreateAPIView):
    serializer_class = DashboardClientSerializer

    def get_queryset(self):
        if getattr(self, "swagger_fake_view", False):
            return Client.objects.none()
        queryset = Client.objects.select_related("current_tariff").annotate(addresses_count=Count("addresses", distinct=True))
        search = self.request.query_params.get("search")
        is_active = bool_param(self.request.query_params.get("is_active"))
        if search:
            queryset = queryset.filter(Q(first_name__icontains=search) | Q(last_name__icontains=search) | Q(phone__icontains=search))
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active)
        return queryset.order_by("-created_at")


@extend_schema(
    tags=[DASHBOARD_CLIENTS_TAG],
    summary="Dashboard client detail/update/delete",
    description="Mijoz detail drawer/modal uchun bitta client ma'lumotini ko'rish, tahrirlash yoki o'chirish uchun ishlatiladi.",
)
class DashboardClientDetailAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.RetrieveUpdateDestroyAPIView):
    serializer_class = DashboardClientSerializer
    queryset = Client.objects.select_related("current_tariff").all()


@extend_schema(
    tags=[DASHBOARD_CLIENTS_TAG],
    summary="Client orders",
    description="Tanlangan mijozga tegishli barcha buyurtmalar tarixini jadvalda ko'rsatish uchun ishlatiladi.",
)
class DashboardClientOrdersAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.ListAPIView):
    serializer_class = DashboardOrderSerializer

    def get_queryset(self):
        if getattr(self, "swagger_fake_view", False):
            return Order.objects.none()
        queryset = (
            Order.objects.filter(client_id=self.kwargs["pk"])
            .select_related("client", "master", "service", "service__category", "address", "tracking")
            .order_by("-created_at")
        )
        return filter_by_category(queryset, self.request, field="service__category")


@extend_schema(
    tags=[DASHBOARD_MASTERS_TAG],
    summary="Dashboard masters list/create",
    description="Ustalar sahifasi uchun ustalar ro'yxatini search/status/specialization filter bilan qaytaradi yoki yangi usta yaratadi.",
    parameters=[
        OpenApiParameter("search", str, OpenApiParameter.QUERY, required=False, description="Ism, familiya yoki telefon bo'yicha qidiruv."),
        OpenApiParameter("status", str, OpenApiParameter.QUERY, required=False, description="active, busy, inactive, blocked."),
        OpenApiParameter(
            "specialization",
            str,
            OpenApiParameter.QUERY,
            required=False,
            description="Mutaxassislik bo'yicha filter. Vergul bilan bir nechta qiymat berish mumkin (masalan `Santexnik,Elektrik`).",
        ),
    ],
)
class DashboardMasterListCreateAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.ListCreateAPIView):
    serializer_class = DashboardMasterSerializer

    def get_queryset(self):
        if getattr(self, "swagger_fake_view", False):
            return Master.objects.none()
        queryset = Master.objects.annotate(
            orders_count=Count("orders", distinct=True),
            completed_orders_count=Count("orders", filter=Q(orders__status=OrderStatus.COMPLETED), distinct=True),
        )
        search = self.request.query_params.get("search")
        status = self.request.query_params.get("status")
        if search:
            queryset = queryset.filter(Q(first_name__icontains=search) | Q(last_name__icontains=search) | Q(phone__icontains=search))
        queryset = filter_masters_by_specialization(queryset, self.request.query_params.get("specialization"))
        if status == "active":
            queryset = queryset.filter(is_blocked=False, is_active=True, is_online=True, is_available=True)
        elif status == "busy":
            queryset = queryset.filter(is_blocked=False, is_active=True, is_online=True, is_available=False)
        elif status == "inactive":
            queryset = queryset.filter(is_blocked=False, is_active=True, is_online=False)
        elif status == "blocked":
            queryset = queryset.filter(is_blocked=True)
        return queryset.order_by("-created_at")


@extend_schema(
    tags=[DASHBOARD_MASTERS_TAG],
    summary="Dashboard master detail/update/delete",
    description="Usta profil/detail oynasi uchun bitta ustani ko'rish, ma'lumotlarini yangilash yoki o'chirish uchun ishlatiladi.",
)
class DashboardMasterDetailAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.RetrieveUpdateDestroyAPIView):
    serializer_class = DashboardMasterSerializer
    queryset = Master.objects.all()


@extend_schema(
    tags=[DASHBOARD_MASTERS_TAG],
    summary="Available masters",
    description="Buyurtmaga usta biriktirish modalida ko'rsatish uchun online va available ustalar ro'yxatini qaytaradi.",
)
class DashboardAvailableMastersAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.ListAPIView):
    serializer_class = DashboardMasterSerializer

    def get_queryset(self):
        return Master.objects.filter(
            is_blocked=False, is_active=True, is_online=True, is_available=True
        ).order_by("first_name")


@extend_schema(
    tags=[DASHBOARD_MASTERS_TAG],
    summary="Bloklangan ustalar",
    description="Bloklangan (is_blocked=True) ustalar ro'yxati. `search` va `specialization` bilan filtrlash mumkin.",
    parameters=[
        OpenApiParameter("search", str, OpenApiParameter.QUERY, required=False, description="Ism, familiya yoki telefon bo'yicha qidiruv."),
        OpenApiParameter("specialization", str, OpenApiParameter.QUERY, required=False, description="Mutaxassislik bo'yicha filter."),
    ],
)
class DashboardBlockedMasterListAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.ListAPIView):
    serializer_class = DashboardMasterSerializer

    def get_queryset(self):
        if getattr(self, "swagger_fake_view", False):
            return Master.objects.none()
        queryset = Master.objects.filter(is_blocked=True)
        search = self.request.query_params.get("search")
        if search:
            queryset = queryset.filter(
                Q(first_name__icontains=search) | Q(last_name__icontains=search) | Q(phone__icontains=search)
            )
        queryset = filter_masters_by_specialization(queryset, self.request.query_params.get("specialization"))
        return queryset.order_by("-blocked_at", "-created_at")


@extend_schema(
    tags=[DASHBOARD_MASTERS_TAG],
    summary="Ustani bloklash / blokdan chiqarish",
    description="`is_blocked=true` ustani bloklaydi (kirishga ruxsat berilmaydi), `false` blokdan chiqaradi. `reason` ixtiyoriy.",
)
class DashboardMasterBlockAPIView(DashboardPermissionMixin, generics.GenericAPIView):
    serializer_class = DashboardMasterBlockSerializer

    def post(self, request, pk):
        master = get_object_or_404(Master, pk=pk)
        serializer = self.get_serializer(data=request.data, context={"master": master})
        serializer.is_valid(raise_exception=True)
        master = serializer.save()
        return success_response(DashboardMasterSerializer(master, context={"request": request}).data)


@extend_schema(
    tags=[DASHBOARD_FINANCE_TAG],
    summary="Masterdan naqd pul qabul qilish (ro'yxat)",
    description="Ustalar topshirgan naqd pul so'rovlari. Default `status=pending`. `status` bilan filtrlash mumkin (pending/approved/rejected).",
    parameters=[
        OpenApiParameter("status", str, OpenApiParameter.QUERY, required=False, description="pending (default), approved yoki rejected."),
    ],
)
class DashboardCashHandoverListAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.ListAPIView):
    serializer_class = DashboardCashHandoverSerializer

    def get_queryset(self):
        if getattr(self, "swagger_fake_view", False):
            return WithdrawRequest.objects.none()
        valid_statuses = {choice[0] for choice in WithdrawRequest.STATUSES}
        status = self.request.query_params.get("status")
        if status not in valid_statuses:
            status = WithdrawRequest.PENDING
        return WithdrawRequest.objects.filter(status=status).select_related("master").order_by("-created_at")


@extend_schema(
    tags=[DASHBOARD_FINANCE_TAG],
    summary="Naqd pulni qabul qilish (tasdiqlash)",
    description=(
        "Admin ustadan naqd pulni qabul qiladi: master naqd balansidan summa yechiladi, OUT tranzaksiya "
        "yoziladi va so'rov `approved` holatiga o'tadi. So'rov `pending` bo'lmasa balans o'zgarmaydi "
        "(idempotent), javobda so'rovning joriy holati qaytadi."
    ),
    request=DashboardCashHandoverActionSerializer,
    responses={200: DashboardCashHandoverEnvelopeSerializer},
    examples=[
        OpenApiExample("So'rov (izoh bilan)", value={"note": "Naqd to'liq qabul qilindi"}, request_only=True),
        OpenApiExample("So'rov (izohsiz)", value={}, request_only=True),
        OpenApiExample(
            "Javob",
            value={
                "success": True,
                "message": "OK",
                "data": {
                    "id": "8f1d3c2a-1b2c-4d5e-8a9b-0c1d2e3f4a5b",
                    "master": "3a2b1c0d-9e8f-7a6b-5c4d-3e2f1a0b9c8d",
                    "master_detail": {
                        "id": "3a2b1c0d-9e8f-7a6b-5c4d-3e2f1a0b9c8d",
                        "full_name": "Jasur Karimov",
                        "phone": "+998901234567",
                        "avatar": None,
                        "rating": "5.00",
                        "is_online": True,
                        "is_available": True,
                    },
                    "amount": "488000.00",
                    "status": "approved",
                    "status_label": "Approved",
                    "admin_note": "Naqd to'liq qabul qilindi",
                    "created_at": "2026-04-23T12:33:00Z",
                    "updated_at": "2026-04-23T12:40:00Z",
                },
            },
            response_only=True,
            status_codes=["200"],
        ),
    ],
)
class DashboardCashHandoverAcceptAPIView(DashboardPermissionMixin, generics.GenericAPIView):
    serializer_class = DashboardCashHandoverActionSerializer

    def post(self, request, pk):
        handover = get_object_or_404(WithdrawRequest, pk=pk)
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        accept_cash_handover(handover, note=serializer.validated_data.get("note", ""))
        handover.refresh_from_db()
        return success_response(DashboardCashHandoverSerializer(handover, context={"request": request}).data)


@extend_schema(
    tags=[DASHBOARD_FINANCE_TAG],
    summary="Naqd pul so'rovini rad etish",
    description=(
        "Admin naqd topshirish so'rovini rad etadi: balans o'zgarmaydi, so'rov `rejected` holatiga o'tadi. "
        "So'rov `pending` bo'lmasa holat o'zgarmaydi (idempotent)."
    ),
    request=DashboardCashHandoverActionSerializer,
    responses={200: DashboardCashHandoverEnvelopeSerializer},
    examples=[
        OpenApiExample("So'rov", value={"note": "Hujjat yetarli emas"}, request_only=True),
        OpenApiExample(
            "Javob",
            value={
                "success": True,
                "message": "OK",
                "data": {
                    "id": "8f1d3c2a-1b2c-4d5e-8a9b-0c1d2e3f4a5b",
                    "master": "3a2b1c0d-9e8f-7a6b-5c4d-3e2f1a0b9c8d",
                    "master_detail": {
                        "id": "3a2b1c0d-9e8f-7a6b-5c4d-3e2f1a0b9c8d",
                        "full_name": "Jasur Karimov",
                        "phone": "+998901234567",
                        "avatar": None,
                        "rating": "5.00",
                        "is_online": True,
                        "is_available": True,
                    },
                    "amount": "488000.00",
                    "status": "rejected",
                    "status_label": "Rejected",
                    "admin_note": "Hujjat yetarli emas",
                    "created_at": "2026-04-23T12:33:00Z",
                    "updated_at": "2026-04-23T12:40:00Z",
                },
            },
            response_only=True,
        ),
    ],
)
class DashboardCashHandoverRejectAPIView(DashboardPermissionMixin, generics.GenericAPIView):
    serializer_class = DashboardCashHandoverActionSerializer

    def post(self, request, pk):
        handover = get_object_or_404(WithdrawRequest, pk=pk)
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        reject_cash_handover(handover, note=serializer.validated_data.get("note", ""))
        handover.refresh_from_db()
        return success_response(DashboardCashHandoverSerializer(handover, context={"request": request}).data)


@extend_schema(
    tags=[DASHBOARD_MASTERS_TAG],
    summary="Master specializations",
    description="Ustalar ro'yxatidagi specialization filter dropdowni uchun mavjud mutaxassisliklar va ular bo'yicha ustalar sonini qaytaradi.",
)
class DashboardMasterSpecializationsAPIView(DashboardPermissionMixin, generics.GenericAPIView):
    serializer_class = EmptySerializer

    def get(self, request):
        results = specialization_counts()
        return success_response({"count": len(results), "results": results})


@extend_schema(
    tags=[DASHBOARD_MASTERS_TAG],
    summary="Master applications (ariza qoldirgan / active bo'lmagan)",
    description=(
        "Ro'yxatdan o'tish arizasini qoldirgan, hali tasdiqlanmagan (active bo'lmagan) ustalar ro'yxati. "
        "Default `approval_status=pending`. `approval_status` (pending/rejected/approved), `search` va "
        "`specialization` bilan filtrlash mumkin."
    ),
    parameters=[
        OpenApiParameter("approval_status", str, OpenApiParameter.QUERY, required=False, description="pending (default), rejected yoki approved."),
        OpenApiParameter("search", str, OpenApiParameter.QUERY, required=False, description="Ism, familiya yoki telefon bo'yicha qidiruv."),
        OpenApiParameter("specialization", str, OpenApiParameter.QUERY, required=False, description="Mutaxassislik bo'yicha filter (vergul bilan bir nechta)."),
    ],
)
class DashboardMasterApplicationListAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.ListAPIView):
    serializer_class = DashboardMasterApplicationSerializer

    def get_queryset(self):
        if getattr(self, "swagger_fake_view", False):
            return Master.objects.none()
        valid_statuses = {choice[0] for choice in MasterApprovalStatus.choices}
        approval_status = self.request.query_params.get("approval_status")
        if approval_status not in valid_statuses:
            approval_status = MasterApprovalStatus.PENDING
        queryset = Master.objects.filter(approval_status=approval_status)
        search = self.request.query_params.get("search")
        if search:
            queryset = queryset.filter(
                Q(first_name__icontains=search) | Q(last_name__icontains=search) | Q(phone__icontains=search)
            )
        queryset = filter_masters_by_specialization(queryset, self.request.query_params.get("specialization"))
        return queryset.order_by("-created_at")


@extend_schema(
    tags=[DASHBOARD_MASTERS_TAG],
    summary="Update master location",
    description="Admin dashboarddan usta lokatsiyasi va online/available holatini yangilash uchun ishlatiladi.",
)
class DashboardMasterLocationAPIView(DashboardPermissionMixin, generics.GenericAPIView):
    serializer_class = DashboardMasterLocationSerializer

    def patch(self, request, pk):
        master = get_object_or_404(Master, pk=pk)
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        master = serializer.save(master=master)
        return success_response(DashboardMasterSerializer(master, context={"request": request}).data)


@extend_schema(
    tags=[DASHBOARD_MASTERS_TAG],
    summary="Update master status",
    description="Usta statusini active, busy, inactive yoki blocked holatlariga o'tkazish uchun ishlatiladi.",
)
class DashboardMasterStatusAPIView(DashboardPermissionMixin, generics.GenericAPIView):
    serializer_class = DashboardMasterStatusSerializer

    def patch(self, request, pk):
        master = get_object_or_404(Master, pk=pk)
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        master = serializer.save(master=master)
        return success_response(DashboardMasterSerializer(master, context={"request": request}).data)


@extend_schema(
    tags=[DASHBOARD_MASTERS_TAG],
    summary="Master orders",
    description="Tanlangan ustaga biriktirilgan buyurtmalar tarixini ko'rsatish uchun ishlatiladi.",
)
class DashboardMasterOrdersAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.ListAPIView):
    serializer_class = DashboardOrderSerializer

    def get_queryset(self):
        if getattr(self, "swagger_fake_view", False):
            return Order.objects.none()
        queryset = filter_by_category(
            Order.objects.filter(master_id=self.kwargs["pk"]),
            self.request,
            field="service__category",
        )
        return (
            queryset
            .select_related("client", "master", "service", "service__category", "address", "tracking")
            .order_by("-created_at")
        )


@extend_schema(
    tags=[DASHBOARD_ORDERS_TAG],
    summary="Dashboard orders list/create",
    description="Buyurtmalar sahifasi uchun orderlar ro'yxatini search/status/payment/date filterlar bilan qaytaradi yoki dashboarddan yangi order yaratadi.",
)
class DashboardOrderListCreateAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.ListCreateAPIView):
    serializer_class = DashboardOrderSerializer

    def get_queryset(self):
        if getattr(self, "swagger_fake_view", False):
            return Order.objects.none()
        queryset = Order.objects.select_related(
            "client", "master", "service", "service__category", "address", "tracking"
        ).prefetch_related("assigned_masters__master", "dashboard_assistants__assistant")
        search = self.request.query_params.get("search")
        status = self.request.query_params.get("status")
        tab = self.request.query_params.get("tab")
        payment_type = self.request.query_params.get("payment_type")
        client = self.request.query_params.get("client")
        master = self.request.query_params.get("master")
        service = self.request.query_params.get("service")
        date_from = self.request.query_params.get("date_from")
        date_to = self.request.query_params.get("date_to")
        if search:
            query = Q(address_text__icontains=search) | Q(note__icontains=search) | Q(client__phone__icontains=search)
            query |= Q(client__first_name__icontains=search) | Q(client__last_name__icontains=search)
            query |= Q(master__first_name__icontains=search) | Q(master__last_name__icontains=search)
            query |= Q(service__name__icontains=search)
            query_uuid = parse_uuid(search)
            if query_uuid:
                query |= Q(id=query_uuid)
            queryset = queryset.filter(query)
        # Figma status tab (Yangi/Yo'lda/Bajarilmoqda/Yakunlangan/Bekor).
        queryset = apply_order_tab(queryset, tab)
        if status:
            queryset = queryset.filter(status=status)
        if payment_type:
            queryset = queryset.filter(payment_type=payment_type)
        if client:
            queryset = queryset.filter(client_id=client)
        if master:
            # Match the lead master OR an assigned master.
            queryset = queryset.filter(
                Q(master_id=master) | Q(assigned_masters__master_id=master, assigned_masters__is_active=True)
            ).distinct()
        if service:
            queryset = queryset.filter(service_id=service)
        queryset = filter_by_category(queryset, self.request, field="service__category")
        if date_from:
            queryset = queryset.filter(created_at__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(created_at__date__lte=date_to)
        return queryset.order_by("-created_at")


@extend_schema(
    tags=[DASHBOARD_ORDERS_TAG],
    summary="Dashboard order detail/update/delete",
    description="Buyurtma detail oynasi uchun bitta orderni ko'rish, tahrirlash yoki o'chirish uchun ishlatiladi.",
)
class DashboardOrderDetailAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.RetrieveUpdateDestroyAPIView):
    serializer_class = DashboardOrderSerializer
    queryset = Order.objects.select_related("client", "master", "service", "service__category", "address", "tracking").all()


@extend_schema(
    tags=[DASHBOARD_ORDERS_TAG],
    summary="Dashboard order board",
    description="Buyurtmalarni status ustunlari bo'yicha board/kanban ko'rinishida chiqarish uchun statuslarga ajratilgan ro'yxat qaytaradi.",
)
class DashboardOrderBoardAPIView(DashboardPermissionMixin, generics.GenericAPIView):
    serializer_class = EmptySerializer

    def get(self, request):
        limit = int(request.query_params.get("limit", 10))
        data = []
        for value, label in OrderStatus.choices:
            queryset = (
                Order.objects.filter(status=value)
                .select_related("client", "master", "service", "service__category", "address", "tracking")
                .order_by("-created_at")
            )
            queryset = filter_by_category(queryset, request, field="service__category")
            data.append(
                {
                    "status": value,
                    "label": label,
                    "count": queryset.count(),
                    "results": DashboardOrderSerializer(queryset[:limit], many=True, context={"request": request}).data,
                }
            )
        return success_response({"columns": data})


@extend_schema(
    tags=[DASHBOARD_ORDERS_TAG],
    summary="Update order status",
    description="Buyurtma statusini yangi, qabul qilingan, jarayonda, yakunlangan, bekor qilingan yoki rad etilgan holatga o'zgartirish uchun ishlatiladi.",
)
class DashboardOrderStatusAPIView(DashboardPermissionMixin, generics.GenericAPIView):
    serializer_class = DashboardOrderStatusSerializer

    def patch(self, request, pk):
        order = get_object_or_404(Order, pk=pk)
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        order = serializer.save(order=order)
        return success_response(DashboardOrderSerializer(order, context={"request": request}).data)


@extend_schema(
    tags=[DASHBOARD_ORDERS_TAG],
    summary="Assign order master",
    description="Usta biriktirish modalidan orderga master biriktirish yoki master qiymatini yangilash uchun ishlatiladi.",
)
class DashboardOrderAssignAPIView(DashboardPermissionMixin, generics.GenericAPIView):
    serializer_class = DashboardOrderAssignSerializer

    def patch(self, request, pk):
        order = get_object_or_404(Order, pk=pk)
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        order = serializer.save(order=order)
        return success_response(DashboardOrderSerializer(order, context={"request": request}).data)


@extend_schema(
    tags=[DASHBOARD_ORDERS_TAG],
    summary="Order tracking snapshot",
    description="Buyurtma tracking oynasi uchun usta lokatsiyasi, masofa, ETA va tracking step ma'lumotlarini qaytaradi.",
)
class DashboardOrderTrackingAPIView(DashboardPermissionMixin, generics.GenericAPIView):
    serializer_class = EmptySerializer

    def get(self, request, pk):
        order = get_object_or_404(Order.objects.select_related("tracking", "master"), pk=pk)
        return success_response(DashboardOrderSerializer(order, context={"request": request}).data["tracking"])


@extend_schema(
    tags=[DASHBOARD_SERVICES_TAG],
    summary="Service categories list/create",
    description="Xizmatlar va Narxlar sahifasi uchun service category ro'yxatini qaytaradi yoki yangi category yaratadi.",
)
class DashboardServiceCategoryListCreateAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.ListCreateAPIView):
    serializer_class = DashboardServiceCategorySerializer

    def get_queryset(self):
        if getattr(self, "swagger_fake_view", False):
            return ServiceCategory.objects.none()
        queryset = ServiceCategory.objects.annotate(services_count=Count("services", distinct=True))
        search = self.request.query_params.get("search")
        is_active = bool_param(self.request.query_params.get("is_active"))
        if search:
            queryset = queryset.filter(Q(name__icontains=search) | Q(slug__icontains=search))
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active)
        return queryset.order_by("sort_order", "name")


@extend_schema(
    tags=[DASHBOARD_SERVICES_TAG],
    summary="Service category detail/update/delete",
    description="Service category detail/edit modal uchun categoryni ko'rish, tahrirlash yoki o'chirish uchun ishlatiladi.",
)
class DashboardServiceCategoryDetailAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.RetrieveUpdateDestroyAPIView):
    serializer_class = DashboardServiceCategorySerializer
    queryset = ServiceCategory.objects.all()


@extend_schema(
    tags=[DASHBOARD_SERVICES_TAG],
    summary="Services list/create",
    description="Xizmatlar jadvali/kartalari uchun service ro'yxatini qaytaradi yoki yangi service yaratadi.",
)
class DashboardServiceListCreateAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.ListCreateAPIView):
    serializer_class = DashboardServiceSerializer

    def get_queryset(self):
        if getattr(self, "swagger_fake_view", False):
            return Service.objects.none()
        queryset = Service.objects.select_related("category").prefetch_related("prices")
        search = self.request.query_params.get("search")
        category = self.request.query_params.get("category")
        is_active = bool_param(self.request.query_params.get("is_active"))
        if search:
            queryset = queryset.filter(Q(name__icontains=search) | Q(description__icontains=search))
        if category:
            queryset = queryset.filter(category_id=category)
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active)
        return queryset.order_by("category__sort_order", "name")


@extend_schema(
    tags=[DASHBOARD_SERVICES_TAG],
    summary="Service detail/update/delete",
    description="Service detail/edit modal uchun xizmatni ko'rish, yangilash yoki o'chirish uchun ishlatiladi.",
)
class DashboardServiceDetailAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.RetrieveUpdateDestroyAPIView):
    serializer_class = DashboardServiceSerializer
    queryset = Service.objects.select_related("category").prefetch_related("prices").all()


@extend_schema(
    tags=[DASHBOARD_TARIFFS_TAG],
    summary="Tariffs list/create",
    description="Tariflar sahifasidagi tarif kartalari uchun tariflar ro'yxatini qaytaradi yoki yangi tarif yaratadi.",
)
class DashboardTariffListCreateAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.ListCreateAPIView):
    serializer_class = DashboardTariffSerializer

    def get_queryset(self):
        queryset = Tariff.objects.prefetch_related("features").annotate(clients_count=Count("clients", distinct=True))
        is_active = bool_param(self.request.query_params.get("is_active"))
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active)
        return queryset.order_by("sort_order", "price", "name")


@extend_schema(
    tags=[DASHBOARD_TARIFFS_TAG],
    summary="Tariff detail/update/delete",
    description="Tarif detail/edit modal uchun tarifni ko'rish, tahrirlash yoki o'chirish uchun ishlatiladi.",
)
class DashboardTariffDetailAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.RetrieveUpdateDestroyAPIView):
    serializer_class = DashboardTariffSerializer
    queryset = Tariff.objects.prefetch_related("features").all()


@extend_schema(
    tags=[DASHBOARD_NOTIFICATIONS_TAG],
    summary="Notifications list/create",
    description="Dashboard notification listi uchun bildirishnomalarni qaytaradi yoki client/masterga yangi notification yaratadi.",
)
class DashboardNotificationListCreateAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.ListCreateAPIView):
    serializer_class = DashboardNotificationSerializer

    def get_queryset(self):
        queryset = Notification.objects.select_related("client", "master")
        role = self.request.query_params.get("role")
        is_read = bool_param(self.request.query_params.get("is_read"))
        if role:
            queryset = queryset.filter(role=role)
        if is_read is not None:
            queryset = queryset.filter(is_read=is_read)
        return queryset.order_by("-created_at")

    def perform_create(self, serializer):
        notification = serializer.save()
        broadcast_notification(notification)
        send_push_notification(notification)


@extend_schema(
    tags=[DASHBOARD_NOTIFICATIONS_TAG],
    summary="Notification detail/update/delete",
    description="Bitta notificationni ko'rish, tahrirlash yoki o'chirish uchun ishlatiladi.",
)
class DashboardNotificationDetailAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.RetrieveUpdateDestroyAPIView):
    serializer_class = DashboardNotificationSerializer
    queryset = Notification.objects.select_related("client", "master").all()


@extend_schema(
    tags=[DASHBOARD_NOTIFICATIONS_TAG],
    summary="Notification unread count",
    description="Header/bell icon uchun o'qilmagan notificationlar sonini qaytaradi.",
)
class DashboardNotificationUnreadCountAPIView(DashboardPermissionMixin, generics.GenericAPIView):
    serializer_class = EmptySerializer

    def get(self, request):
        return success_response({"unread_count": Notification.objects.filter(is_read=False).count()})


@extend_schema(
    tags=[DASHBOARD_NOTIFICATIONS_TAG],
    summary="Mark notification read",
    description="Tanlangan notificationni o'qilgan holatga o'tkazadi.",
)
class DashboardNotificationReadAPIView(DashboardPermissionMixin, generics.GenericAPIView):
    serializer_class = EmptySerializer

    def post(self, request, pk):
        notification = get_object_or_404(Notification, pk=pk)
        notification.is_read = True
        notification.save(update_fields=["is_read", "updated_at"])
        return success_response(DashboardNotificationSerializer(notification, context={"request": request}).data)


@extend_schema(
    tags=[DASHBOARD_NOTIFICATIONS_TAG],
    summary="Mark all notifications read",
    description="Dashboarddagi barcha o'qilmagan notificationlarni o'qilgan holatga o'tkazadi.",
)
class DashboardNotificationReadAllAPIView(DashboardPermissionMixin, generics.GenericAPIView):
    serializer_class = EmptySerializer

    def post(self, request):
        Notification.objects.filter(is_read=False).update(is_read=True)
        return success_response(message="All notifications marked as read")


@extend_schema(
    tags=[DASHBOARD_EXPENSES_TAG],
    summary="Expenses list/create",
    description="Xarajatlar sahifasidagi 'Usta xarajatlari' tabi uchun master expense ro'yxatini qaytaradi yoki yangi xarajat yaratadi.",
)
class DashboardExpenseListCreateAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.ListCreateAPIView):
    serializer_class = DashboardExpenseSerializer

    def get_queryset(self):
        queryset = MasterExpense.objects.select_related("master")
        master = self.request.query_params.get("master")
        date_from = self.request.query_params.get("date_from")
        date_to = self.request.query_params.get("date_to")
        if master:
            queryset = queryset.filter(master_id=master)
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)
        return queryset.order_by("-date", "-created_at")


@extend_schema(
    tags=[DASHBOARD_EXPENSES_TAG],
    summary="Expense detail/update/delete",
    description="Usta xarajati detail/edit modalida xarajatni ko'rish, tahrirlash yoki o'chirish uchun ishlatiladi.",
)
class DashboardExpenseDetailAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.RetrieveUpdateDestroyAPIView):
    serializer_class = DashboardExpenseSerializer
    queryset = MasterExpense.objects.select_related("master").all()


@extend_schema(
    tags=[DASHBOARD_STAFF_TAG],
    summary="Dashboard staff list/create",
    description="Xodim sahifasi uchun dashboard staff/admin userlar ro'yxatini qaytaradi yoki yangi xodim yaratadi.",
)
class DashboardStaffListCreateAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.ListCreateAPIView):
    serializer_class = DashboardStaffSerializer

    def get_queryset(self):
        user_model = get_user_model()
        queryset = user_model.objects.filter(is_staff=True).select_related("dashboard_profile")
        search = self.request.query_params.get("search")
        role = self.request.query_params.get("role")
        is_active = bool_param(self.request.query_params.get("is_active"))
        if search:
            queryset = queryset.filter(
                Q(username__icontains=search)
                | Q(first_name__icontains=search)
                | Q(last_name__icontains=search)
                | Q(email__icontains=search)
                | Q(dashboard_profile__phone__icontains=search)
            )
        if role:
            queryset = queryset.filter(dashboard_profile__role=role)
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active)
        return queryset.order_by("username")


@extend_schema(
    tags=[DASHBOARD_STAFF_TAG],
    summary="Dashboard staff detail/update/delete",
    description="Xodim detail/edit modal uchun staff user profilini ko'rish, yangilash yoki o'chirish uchun ishlatiladi.",
)
class DashboardStaffDetailAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.RetrieveUpdateDestroyAPIView):
    serializer_class = DashboardStaffSerializer

    def get_queryset(self):
        return get_user_model().objects.filter(is_staff=True).select_related("dashboard_profile")


@extend_schema(
    tags=[DASHBOARD_MAPS_TAG],
    summary="Map master markers",
    description="Xaritalar sahifasi uchun lokatsiyasi bor ustalarni marker ma'lumotlari bilan qaytaradi.",
)
class DashboardMapMastersAPIView(DashboardPermissionMixin, generics.GenericAPIView):
    serializer_class = EmptySerializer

    def get(self, request):
        queryset = Master.objects.filter(lat__isnull=False, lng__isnull=False)
        status = request.query_params.get("status")
        if status == "active":
            queryset = queryset.filter(is_active=True, is_online=True, is_available=True)
        elif status == "busy":
            queryset = queryset.filter(is_active=True, is_online=True, is_available=False)
        elif status == "offline":
            queryset = queryset.filter(is_active=True, is_online=False)
        elif status == "blocked":
            queryset = queryset.filter(is_active=False)
        data = [
            {
                "id": master.id,
                "full_name": master.full_name,
                "phone": master.phone,
                "specialization": master.specialization,
                "avatar": master.avatar.url if master.avatar else None,
                "rating": master.rating,
                "lat": master.lat,
                "lng": master.lng,
                "is_online": master.is_online,
                "is_available": master.is_available,
                "is_active": master.is_active,
                "last_location_at": master.last_location_at,
                "status": DashboardMasterSerializer().get_status(master),
            }
            for master in queryset.order_by("first_name", "last_name")
        ]
        return success_response({"count": len(data), "results": data})


@extend_schema(
    tags=[DASHBOARD_CLIENTS_TAG],
    summary="Client stats summary",
    description="Mijoz detail sahifasi uchun orderlar soni, sarflangan summa, address/device/market statistikalarini qaytaradi.",
)
class DashboardClientStatsAPIView(DashboardPermissionMixin, generics.GenericAPIView):
    serializer_class = EmptySerializer

    def get(self, request, pk):
        client = get_object_or_404(Client.objects.select_related("current_tariff"), pk=pk)
        orders = Order.objects.filter(client=client)
        completed = orders.filter(status=OrderStatus.COMPLETED)
        data = {
            "client": DashboardClientSerializer(client, context={"request": request}).data,
            "orders_count": orders.count(),
            "completed_orders_count": completed.count(),
            "active_orders_count": orders.filter(status__in=ACTIVE_ORDER_STATUSES).count(),
            "cancelled_orders_count": orders.filter(status=OrderStatus.CANCELLED).count(),
            "total_spent": completed.aggregate(total=Sum("total_amount"))["total"] or 0,
            "last_order_at": orders.aggregate(last=Max("created_at"))["last"],
            "addresses_count": client.addresses.count(),
            "devices_count": client.client_devices.count(),
            "market_orders_count": client.market_orders.count(),
        }
        return success_response(data)


@extend_schema(
    tags=[DASHBOARD_ORDERS_TAG],
    summary="Order assistant list/create",
    description="Shogird biriktirish modalida orderga assistant master biriktirish yoki mavjud assistantlarni ko'rish uchun ishlatiladi.",
)
class DashboardOrderAssistantListCreateAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.ListCreateAPIView):
    serializer_class = DashboardOrderAssistantSerializer

    def get_queryset(self):
        if getattr(self, "swagger_fake_view", False):
            return DashboardOrderAssistant.objects.none()
        return DashboardOrderAssistant.objects.filter(order_id=self.kwargs["pk"]).select_related("assistant", "assigned_by")

    def perform_create(self, serializer):
        order = get_object_or_404(Order, pk=self.kwargs["pk"])
        serializer.save(order=order, assigned_by=self.request.user)


@extend_schema(
    tags=[DASHBOARD_ORDERS_TAG],
    summary="Order assistant detail/update/delete",
    description="Orderga biriktirilgan assistant master yozuvini ko'rish, yangilash yoki olib tashlash uchun ishlatiladi.",
)
class DashboardOrderAssistantDetailAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.RetrieveUpdateDestroyAPIView):
    serializer_class = DashboardOrderAssistantSerializer
    lookup_url_kwarg = "assistant_pk"

    def get_queryset(self):
        return DashboardOrderAssistant.objects.filter(order_id=self.kwargs["pk"]).select_related("assistant", "assigned_by")


@extend_schema(
    tags=[DASHBOARD_SERVICES_TAG],
    summary="Service price list/create",
    description="Tanlangan service uchun narx variantlari ro'yxatini qaytaradi yoki yangi narx qo'shadi.",
)
class DashboardServicePriceListCreateAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.ListCreateAPIView):
    serializer_class = DashboardServicePriceSerializer

    def get_queryset(self):
        if getattr(self, "swagger_fake_view", False):
            return ServicePrice.objects.none()
        return ServicePrice.objects.filter(service_id=self.kwargs["service_id"]).order_by("title")

    def perform_create(self, serializer):
        service = get_object_or_404(Service, pk=self.kwargs["service_id"])
        serializer.save(service=service)


@extend_schema(
    tags=[DASHBOARD_SERVICES_TAG],
    summary="Service price detail/update/delete",
    description="Service narx variantini ko'rish, tahrirlash yoki o'chirish uchun ishlatiladi.",
)
class DashboardServicePriceDetailAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.RetrieveUpdateDestroyAPIView):
    serializer_class = DashboardServicePriceSerializer
    queryset = ServicePrice.objects.select_related("service").all()


@extend_schema(
    tags=[DASHBOARD_TARIFFS_TAG],
    summary="Tariff feature list/create",
    description="Tanlangan tarifga tegishli afzalliklar/features ro'yxatini qaytaradi yoki yangi feature qo'shadi.",
)
class DashboardTariffFeatureListCreateAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.ListCreateAPIView):
    serializer_class = DashboardTariffFeatureSerializer

    def get_queryset(self):
        if getattr(self, "swagger_fake_view", False):
            return TariffFeature.objects.none()
        return TariffFeature.objects.filter(tariff_id=self.kwargs["tariff_id"]).order_by("sort_order", "id")

    def perform_create(self, serializer):
        tariff = get_object_or_404(Tariff, pk=self.kwargs["tariff_id"])
        serializer.save(tariff=tariff)


@extend_schema(
    tags=[DASHBOARD_TARIFFS_TAG],
    summary="Tariff feature detail/update/delete",
    description="Tarif feature yozuvini ko'rish, tahrirlash yoki o'chirish uchun ishlatiladi.",
)
class DashboardTariffFeatureDetailAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.RetrieveUpdateDestroyAPIView):
    serializer_class = DashboardTariffFeatureSerializer
    queryset = TariffFeature.objects.select_related("tariff").all()


@extend_schema(
    tags=[DASHBOARD_MARKET_TAG],
    summary="Marketplace kategoriyalari",
    description="Marketplace sahifasidagi kategoriya filterlari uchun kategoriyalar ro'yxatini qaytaradi yoki yangi kategoriya yaratadi.",
)
class DashboardMarketCategoryListCreateAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.ListCreateAPIView):
    serializer_class = DashboardMarketCategorySerializer

    def get_queryset(self):
        queryset = MarketCategory.objects.annotate(products_count=Count("marketproduct", distinct=True))
        search = self.request.query_params.get("search")
        if search:
            queryset = queryset.filter(Q(name__icontains=search) | Q(slug__icontains=search))
        return queryset.order_by("name")


@extend_schema(
    tags=[DASHBOARD_MARKET_TAG],
    summary="Marketplace kategoriya detail",
    description="Marketplace kategoriya detail/edit modalida kategoriyani ko'rish, tahrirlash yoki o'chirish uchun ishlatiladi.",
)
class DashboardMarketCategoryDetailAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.RetrieveUpdateDestroyAPIView):
    serializer_class = DashboardMarketCategorySerializer
    queryset = MarketCategory.objects.all()


@extend_schema(
    tags=[DASHBOARD_MARKET_TAG],
    summary="Marketplace mahsulotlari",
    description="Marketplace mahsulotlari sahifasi uchun product ro'yxatini search/filter bilan qaytaradi yoki dashboarddan yangi product yaratadi.",
)
@extend_schema_view(
    get=extend_schema(tags=[DASHBOARD_MARKET_TAG], summary="Marketplace mahsulotlar ro'yxati"),
    post=extend_schema(
        tags=[DASHBOARD_MARKET_TAG],
        summary="Marketplace mahsulot yaratish (rasm bilan birga)",
        description=(
            "Bitta `multipart/form-data` so'rovda mahsulot ma'lumoti VA rasmlar birga yuboriladi. "
            "`uploaded_images` — bir nechta rasm fayli (bir xil nom bilan takrorlanadi). Alohida rasm API shart emas."
        ),
        request={"multipart/form-data": DashboardMarketProductSerializer},
    ),
)
class DashboardMarketProductListCreateAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.ListCreateAPIView):
    serializer_class = DashboardMarketProductSerializer

    def get_queryset(self):
        queryset = MarketProduct.objects.select_related("category", "seller").prefetch_related("images").annotate(
            orders_count=Count("orders", distinct=True)
        )
        search = self.request.query_params.get("search")
        condition = self.request.query_params.get("condition")
        is_active = bool_param(self.request.query_params.get("is_active"))
        is_moderated = bool_param(self.request.query_params.get("is_moderated"))
        category = self.request.query_params.get("category")
        if search:
            queryset = queryset.filter(Q(name__icontains=search) | Q(description__icontains=search))
        if condition:
            queryset = queryset.filter(condition=condition)
        if category:
            queryset = queryset.filter(category_id=category)
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active)
        if is_moderated is not None:
            queryset = queryset.filter(is_moderated=is_moderated)
        return queryset.order_by("-created_at")


@extend_schema_view(
    get=extend_schema(tags=[DASHBOARD_MARKET_TAG], summary="Marketplace mahsulot detail"),
    put=extend_schema(
        tags=[DASHBOARD_MARKET_TAG],
        summary="Mahsulotni tahrirlash (rasm bilan)",
        description="`multipart/form-data`: matn maydonlari + `uploaded_images` (yangi rasmlar qo'shiladi).",
        request={"multipart/form-data": DashboardMarketProductSerializer},
    ),
    patch=extend_schema(
        tags=[DASHBOARD_MARKET_TAG],
        summary="Mahsulotni qisman tahrirlash (rasm bilan)",
        request={"multipart/form-data": DashboardMarketProductSerializer},
    ),
    delete=extend_schema(tags=[DASHBOARD_MARKET_TAG], summary="Mahsulotni o'chirish yoki arxivlash"),
)
class DashboardMarketProductDetailAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.RetrieveUpdateDestroyAPIView):
    serializer_class = DashboardMarketProductSerializer
    queryset = MarketProduct.objects.select_related("category", "seller").prefetch_related("images").all()

    def delete(self, request, *args, **kwargs):
        # MarketOrder.product is on_delete=PROTECT, so a product that already has
        # orders cannot be hard-deleted (it would raise ProtectedError -> 500).
        # Archive it instead (leaves every customer-facing listing) and preserve
        # order history; hard-delete only when there are no orders.
        instance = self.get_object()
        if instance.orders.exists():
            instance.is_active = False
            instance.is_moderated = False
            instance.save(update_fields=["is_active", "is_moderated", "updated_at"])
            return success_response(
                {"id": str(instance.id), "deleted": False, "archived": True},
                message="Mahsulotda buyurtmalar bor — arxivlandi (is_active=False).",
            )
        instance.delete()
        return success_response({"id": str(kwargs.get("pk")), "deleted": True, "archived": False}, message="Mahsulot o'chirildi")


@extend_schema(
    tags=[DASHBOARD_MARKET_TAG],
    summary="Marketplace mahsulot rasmlari",
    description="Tanlangan marketplace mahsulotiga rasm qo'shish yoki mavjud rasmlar ro'yxatini ko'rsatish uchun ishlatiladi.",
)
class DashboardMarketProductImageListCreateAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.ListCreateAPIView):
    serializer_class = DashboardMarketProductImageSerializer

    def get_queryset(self):
        if getattr(self, "swagger_fake_view", False):
            return MarketProductImage.objects.none()
        return MarketProductImage.objects.filter(product_id=self.kwargs["product_id"]).order_by("-created_at")

    def perform_create(self, serializer):
        product = get_object_or_404(MarketProduct, pk=self.kwargs["product_id"])
        serializer.save(product=product)


@extend_schema(
    tags=[DASHBOARD_MARKET_TAG],
    summary="Marketplace mahsulot rasmi detail",
    description="Marketplace mahsulotining bitta rasmini ko'rish yoki o'chirish uchun ishlatiladi.",
)
class DashboardMarketProductImageDetailAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.RetrieveDestroyAPIView):
    serializer_class = DashboardMarketProductImageSerializer
    queryset = MarketProductImage.objects.select_related("product").all()


@extend_schema(
    tags=[DASHBOARD_MARKET_TAG],
    summary="Marketplace buyurtmalari",
    description="Marketplace orderlar jadvali uchun buyurtmalar ro'yxatini status/client/product bo'yicha filterlab qaytaradi yoki yangi market order yaratadi.",
)
class DashboardMarketOrderListCreateAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.ListCreateAPIView):
    serializer_class = DashboardMarketOrderSerializer

    def get_queryset(self):
        queryset = MarketOrder.objects.select_related("client", "product", "product__category", "product__seller")
        status = self.request.query_params.get("status")
        client = self.request.query_params.get("client")
        product = self.request.query_params.get("product")
        if status:
            queryset = queryset.filter(status=status)
        if client:
            queryset = queryset.filter(client_id=client)
        if product:
            queryset = queryset.filter(product_id=product)
        queryset = filter_by_category(queryset, self.request, field="product__category")
        return queryset.order_by("-created_at")


@extend_schema(
    tags=[DASHBOARD_MARKET_TAG],
    summary="Marketplace buyurtma detail",
    description="Marketplace buyurtma detail/edit oynasi uchun market orderni ko'rish, yangilash yoki o'chirish uchun ishlatiladi.",
)
class DashboardMarketOrderDetailAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.RetrieveUpdateDestroyAPIView):
    serializer_class = DashboardMarketOrderSerializer
    queryset = MarketOrder.objects.select_related("client", "product", "product__category", "product__seller").all()


@extend_schema(
    tags=[DASHBOARD_WAREHOUSE_TAG],
    summary="Ombor statistikasi",
    description="Ombor sahifasidagi cardlar uchun mahsulotlar, low stock, jami qoldiq va kirim/chiqim statistikalarini qaytaradi.",
)
class DashboardWarehouseStatsAPIView(DashboardPermissionMixin, generics.GenericAPIView):
    serializer_class = EmptySerializer

    def get(self, request):
        products = WarehouseProduct.objects.all()
        data = {
            "products_count": products.count(),
            "active_products_count": products.filter(is_active=True).count(),
            "low_stock_count": sum(1 for product in products if product.is_low_stock),
            "total_quantity": products.aggregate(total=Sum("quantity"))["total"] or 0,
            "total_value": products.aggregate(
                total=Sum(F("cost_price") * F("quantity"), output_field=DecimalField(max_digits=18, decimal_places=2))
            )["total"] or 0,
            "master_inventory_count": MasterInventory.objects.count(),
            "stock_in_count": StockMovement.objects.filter(movement_type=StockMovement.IN).count(),
            "stock_out_count": StockMovement.objects.filter(movement_type=StockMovement.OUT).count(),
        }
        return success_response(data)


@extend_schema(
    tags=[DASHBOARD_WAREHOUSE_TAG],
    summary="Ombor kategoriyalari",
    description="Ombor mahsulot kategoriyalari ro'yxatini search bilan qaytaradi yoki yangi kategoriya yaratadi.",
)
class DashboardWarehouseCategoryListCreateAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.ListCreateAPIView):
    serializer_class = DashboardWarehouseCategorySerializer

    def get_queryset(self):
        queryset = WarehouseCategory.objects.annotate(products_count=Count("products", distinct=True))
        search = self.request.query_params.get("search")
        if search:
            queryset = queryset.filter(Q(name__icontains=search) | Q(slug__icontains=search))
        return queryset.order_by("name")


@extend_schema(
    tags=[DASHBOARD_WAREHOUSE_TAG],
    summary="Ombor kategoriya detail",
    description="Ombor kategoriya detail/edit modalida kategoriyani ko'rish, tahrirlash yoki o'chirish uchun ishlatiladi.",
)
class DashboardWarehouseCategoryDetailAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.RetrieveUpdateDestroyAPIView):
    serializer_class = DashboardWarehouseCategorySerializer
    queryset = WarehouseCategory.objects.all()


@extend_schema(
    tags=[DASHBOARD_WAREHOUSE_TAG],
    summary="Ombor mahsulotlari",
    description="Ombor mahsulotlari jadvali uchun product ro'yxatini search/active/low_stock/category filterlari bilan qaytaradi yoki yangi ombor mahsuloti yaratadi.",
)
class DashboardWarehouseProductListCreateAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.ListCreateAPIView):
    serializer_class = DashboardWarehouseProductSerializer

    def get_queryset(self):
        queryset = (
            WarehouseProduct.objects.select_related("category")
            .annotate(movements_count=Count("movements", distinct=True))
            .order_by("name")
        )
        search = self.request.query_params.get("search")
        is_active = bool_param(self.request.query_params.get("is_active"))
        low_stock = bool_param(self.request.query_params.get("low_stock"))
        if search:
            queryset = queryset.filter(name__icontains=search)
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active)
        # Applied before the low_stock branch below, which materialises the queryset into a list.
        queryset = filter_by_category(queryset, self.request, field="category")
        if low_stock is True:
            queryset = [product for product in queryset if product.is_low_stock]
        return queryset

    def perform_create(self, serializer):
        # Journal the opening stock so the movement ledger stays in sync with on-hand qty.
        product = serializer.save()
        if product.quantity:
            StockMovement.objects.create(
                product=product,
                movement_type=StockMovement.IN,
                quantity=product.quantity,
                note="Boshlang'ich qoldiq",
            )


@extend_schema(
    tags=[DASHBOARD_WAREHOUSE_TAG],
    summary="Ombor mahsulot detail",
    description="Ombor mahsulot detail/edit modalida mahsulotni ko'rish, yangilash yoki o'chirish uchun ishlatiladi.",
)
class DashboardWarehouseProductDetailAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.RetrieveUpdateDestroyAPIView):
    serializer_class = DashboardWarehouseProductSerializer
    queryset = WarehouseProduct.objects.all()

    def perform_update(self, serializer):
        # A direct `quantity` edit is routed through the locked stock service so it
        # records a StockMovement and can't silently diverge the ledger or go negative.
        product = serializer.instance
        new_quantity = serializer.validated_data.pop("quantity", None)
        if serializer.validated_data:
            serializer.save()
        if new_quantity is not None and new_quantity != product.quantity:
            adjust_warehouse_stock(product, new_quantity - product.quantity, note="Admin qoldiqni to'g'riladi")
        serializer.instance.refresh_from_db()


@extend_schema(
    tags=[DASHBOARD_WAREHOUSE_TAG],
    summary="Ombor kirim-chiqimlari",
    description="Ombor harakatlari jadvali uchun mahsulot kirim/chiqim tarixini qaytaradi yoki yangi stock movement yaratadi.",
)
class DashboardStockMovementListCreateAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.ListCreateAPIView):
    serializer_class = DashboardStockMovementSerializer

    def get_queryset(self):
        queryset = StockMovement.objects.select_related("product", "master")
        product = self.request.query_params.get("product")
        movement_type = self.request.query_params.get("movement_type")
        if product:
            queryset = queryset.filter(product_id=product)
        if movement_type:
            queryset = queryset.filter(movement_type=movement_type)
        return queryset.order_by("-created_at")

    @transaction.atomic
    def perform_create(self, serializer):
        # A manual movement must actually move on-hand stock (IN adds, OUT removes),
        # otherwise the ledger and WarehouseProduct.quantity drift apart.
        data = serializer.validated_data
        apply_stock_movement_effect(data["product"], data["movement_type"], data["quantity"])
        serializer.save()


LEDGER_IMMUTABLE_MSG = "Ledger yozuvining moliyaviy maydonini o'zgartirib bo'lmaydi"


@extend_schema(
    tags=[DASHBOARD_WAREHOUSE_TAG],
    summary="Ombor harakati detail",
    description="Bitta ombor kirim/chiqim yozuvini ko'rish, tahrirlash yoki o'chirish uchun ishlatiladi.",
)
class DashboardStockMovementDetailAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.RetrieveUpdateDestroyAPIView):
    serializer_class = DashboardStockMovementSerializer
    queryset = StockMovement.objects.select_related("product", "master").all()

    def perform_update(self, serializer):
        # Financial fields of a posted movement are immutable (editing them would
        # desync stock); only `note` may be corrected.
        instance = serializer.instance
        for field in ("product", "movement_type", "quantity", "master"):
            if field in serializer.validated_data and serializer.validated_data[field] != getattr(instance, field):
                raise ValidationError({field: LEDGER_IMMUTABLE_MSG})
        serializer.save()

    def perform_destroy(self, instance):
        # Deleting a movement reverses its effect on on-hand stock.
        reverse_stock_movement(instance)
        instance.delete()


@extend_schema(
    tags=[DASHBOARD_WAREHOUSE_TAG],
    summary="Usta inventari",
    description="Ombordan ustalarga biriktirilgan inventarlar ro'yxatini qaytaradi yoki ustaga yangi mahsulot biriktiradi.",
)
class DashboardMasterInventoryListCreateAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.ListCreateAPIView):
    serializer_class = DashboardMasterInventorySerializer

    def get_queryset(self):
        queryset = MasterInventory.objects.select_related("master", "warehouse_product")
        master = self.request.query_params.get("master")
        product = self.request.query_params.get("product")
        low_stock = bool_param(self.request.query_params.get("low_stock"))
        if master:
            queryset = queryset.filter(master_id=master)
        if product:
            queryset = queryset.filter(warehouse_product_id=product)
        if low_stock is True:
            queryset = [item for item in queryset if item.is_low_stock]
        return queryset

    @extend_schema(
        summary="Ustaga mahsulot biriktirish",
        description=(
            "Ustaga mahsulot biriktiradi: markaziy ombordan miqdorni ayiradi, StockMovement (chiqim) "
            "yozadi va usta allaqachon shu mahsulotga ega bo'lsa miqdorni ustiga qo'shadi. Ombor yetarli "
            "bo'lmasa 400 qaytaradi."
        ),
        request=DashboardMasterInventoryAssignSerializer,
        responses={201: DashboardMasterInventorySerializer},
    )
    def create(self, request, *args, **kwargs):
        serializer = DashboardMasterInventoryAssignSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        item = serializer.save()
        data = DashboardMasterInventorySerializer(item, context=self.get_serializer_context()).data
        return success_response(data, status=201)


@extend_schema(
    tags=[DASHBOARD_WAREHOUSE_TAG],
    summary="Usta inventari detail",
    description="Ustaga biriktirilgan bitta inventar yozuvini ko'rish, yangilash yoki o'chirish uchun ishlatiladi.",
)
class DashboardMasterInventoryDetailAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.RetrieveUpdateDestroyAPIView):
    serializer_class = DashboardMasterInventorySerializer
    queryset = MasterInventory.objects.select_related("master", "warehouse_product").all()

    def perform_update(self, serializer):
        # A quantity edit is routed through the locked service (moves the warehouse
        # delta + records a StockMovement); the identity fields can't be reassigned.
        item = serializer.instance
        for field in ("master", "warehouse_product"):
            if field in serializer.validated_data and serializer.validated_data[field] != getattr(item, field):
                raise ValidationError({field: "Bu maydonni o'zgartirib bo'lmaydi"})
        new_quantity = serializer.validated_data.pop("quantity", None)
        if serializer.validated_data:
            serializer.save()
        if new_quantity is not None and new_quantity != item.quantity:
            adjust_master_inventory(item, new_quantity)
        serializer.instance.refresh_from_db()

    def perform_destroy(self, instance):
        # Returns the remaining stock to the central warehouse (+ StockMovement IN).
        return_inventory_to_warehouse(instance)


@extend_schema(
    tags=[DASHBOARD_LIVE_TAG],
    summary="Jonli kuzatuv streamlari",
    description="Jonli kuzatuv sahifasi uchun usta/client/order bo'yicha streamlar ro'yxatini qaytaradi yoki yangi stream yozuvi yaratadi.",
)
class DashboardLiveStreamListCreateAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.ListCreateAPIView):
    serializer_class = DashboardLiveStreamSerializer

    def get_queryset(self):
        queryset = DashboardLiveStream.objects.select_related("master", "client", "order", "order__service")
        status = self.request.query_params.get("status")
        is_active = bool_param(self.request.query_params.get("is_active"))
        if status:
            queryset = queryset.filter(status=status)
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active)
        return queryset.order_by("-created_at")


@extend_schema(
    tags=[DASHBOARD_LIVE_TAG],
    summary="Jonli kuzatuv detail",
    description="Bitta live stream yozuvini ko'rish, status yoki linklarini yangilash, yoki o'chirish uchun ishlatiladi.",
)
class DashboardLiveStreamDetailAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.RetrieveUpdateDestroyAPIView):
    serializer_class = DashboardLiveStreamSerializer
    queryset = DashboardLiveStream.objects.select_related("master", "client", "order", "order__service").all()


@extend_schema(
    tags=[DASHBOARD_LIVE_TAG],
    summary="Arxiv videolar",
    description="Jonli kuzatuv sahifasidagi arxiv blok uchun tugagan yoki arxivlangan stream yozuvlarini qaytaradi.",
)
class DashboardArchivedVideoListAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.ListAPIView):
    serializer_class = DashboardLiveStreamSerializer

    def get_queryset(self):
        return DashboardLiveStream.objects.filter(status__in=[DashboardLiveStream.ARCHIVED, DashboardLiveStream.ENDED]).select_related(
            "master", "client", "order", "order__service"
        )


@extend_schema(
    tags=[DASHBOARD_SUPPORT_TAG],
    summary="Support chat inbox",
    description="Xabarlar sahifasi uchun client/master kesimida support threadlar, oxirgi xabar va unread count ro'yxatini qaytaradi.",
)
class DashboardSupportThreadListAPIView(DashboardPermissionMixin, generics.GenericAPIView):
    serializer_class = EmptySerializer

    def get(self, request):
        chats = (
            SupportChat.objects.select_related("client", "master")
            .annotate(
                unread_count=Count(
                    "messages",
                    filter=Q(messages__is_read=False) & ~Q(messages__sender_role="admin"),
                )
            )
            .order_by("-updated_at")
        )
        attach_latest_support_messages(chats)
        threads = []
        for chat in chats:
            last_message = (
                chat.messages.select_related("client", "master").order_by("-created_at").first()
            )
            if not last_message:
                continue
            has_unread = bool(chat.unread_count)
            threads.append(
                {
                    # chat_id -> WebSocket: ws/support/<chat_id>/
                    "chat_id": str(chat.id),
                    "participant_role": chat.participant_role,
                    "client": DashboardClientMiniSerializer(chat.client).data if chat.client else None,
                    "master": DashboardMasterMiniSerializer(chat.master).data if chat.master else None,
                    "last_message": DashboardSupportMessageSerializer(last_message, context={"request": request}).data,
                    "last_message_at": last_message.created_at,
                    "unread_count": chat.unread_count,
                    "has_unread": has_unread,
                    "status": "unread" if has_unread else "read",
                }
            )
        # Unread threads first, then most recent activity.
        threads.sort(key=lambda thread: (thread["has_unread"], thread["last_message_at"]), reverse=True)
        return success_response({"count": len(threads), "results": threads})


@extend_schema(
    tags=[DASHBOARD_SUPPORT_TAG],
    summary="Support xabarlari",
    description="Tanlangan support thread xabarlarini qaytaradi yoki dashboard admin nomidan yangi support xabar yuboradi.",
)
class DashboardSupportMessageListCreateAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.ListCreateAPIView):
    serializer_class = DashboardSupportMessageSerializer

    def get_queryset(self):
        queryset = SupportMessage.objects.select_related("client", "master")
        client = self.request.query_params.get("client")
        master = self.request.query_params.get("master")
        is_read = bool_param(self.request.query_params.get("is_read"))
        if client:
            queryset = queryset.filter(client_id=client)
        if master:
            queryset = queryset.filter(master_id=master)
        if is_read is not None:
            queryset = queryset.filter(is_read=is_read)
        return queryset.order_by("created_at")

    def list(self, request, *args, **kwargs):
        # Opening a thread counts as the admin reading it: flip is_read on the
        # participant's incoming messages before serializing the response.
        client = request.query_params.get("client")
        master = request.query_params.get("master")
        if client or master:
            mark_support_thread_read_by_admin(client_id=client or None, master_id=master or None)
        return super().list(request, *args, **kwargs)

    def perform_create(self, serializer):
        # The client/master apps read the thread through the ``chat`` FK
        # (SupportChat.messages), so an admin reply saved without it never
        # reaches the participant. Attach the chat and emit the same realtime
        # broadcast the participant-facing endpoints use.
        message = serializer.save(sender_role="admin", admin=self.request.user)
        if not message.chat:
            participant = message.client or message.master
            chat = get_or_create_support_chat(participant) if participant else None
            if chat:
                message.chat = chat
                message.save(update_fields=["chat", "updated_at"])
                touch_chat(chat)
        broadcast_support_message(message)


@extend_schema(
    tags=[DASHBOARD_SUPPORT_TAG],
    summary="Support xabar detail",
    description="Bitta support xabarini ko'rish, tahrirlash yoki o'chirish uchun ishlatiladi.",
)
class DashboardSupportMessageDetailAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.RetrieveUpdateDestroyAPIView):
    serializer_class = DashboardSupportMessageSerializer
    queryset = SupportMessage.objects.select_related("client", "master").all()


@extend_schema(
    tags=[DASHBOARD_SUPPORT_TAG],
    summary="Support xabarni o'qilgan qilish",
    description="Support inboxdagi bitta xabarni o'qilgan holatga o'tkazadi.",
)
class DashboardSupportMessageReadAPIView(DashboardPermissionMixin, generics.GenericAPIView):
    serializer_class = EmptySerializer

    def post(self, request, pk):
        message = get_object_or_404(SupportMessage, pk=pk)
        message.is_read = True
        message.save(update_fields=["is_read", "updated_at"])
        return success_response(DashboardSupportMessageSerializer(message, context={"request": request}).data)


@extend_schema(
    tags=[DASHBOARD_FINANCE_TAG],
    summary="Usta hamyoni",
    description="Usta detail yoki moliya sahifasi uchun tanlangan ustaning wallet balans va umumiy hamyon ma'lumotlarini qaytaradi.",
)
class DashboardMasterWalletAPIView(DashboardPermissionMixin, generics.GenericAPIView):
    serializer_class = EmptySerializer

    def get(self, request, pk):
        master = get_object_or_404(Master, pk=pk)
        wallet, _ = MasterWallet.objects.get_or_create(master=master)
        return success_response(DashboardMasterWalletSerializer(wallet, context={"request": request}).data)


@extend_schema(
    tags=[DASHBOARD_FINANCE_TAG],
    summary="Wallet tranzaksiyalari",
    description="Usta wallet tranzaksiyalari jadvali uchun transaction ro'yxatini qaytaradi yoki yangi wallet transaction yaratadi.",
)
class DashboardWalletTransactionListCreateAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.ListCreateAPIView):
    serializer_class = DashboardWalletTransactionSerializer

    def get_queryset(self):
        queryset = WalletTransaction.objects.select_related("master", "order")
        master = self.request.query_params.get("master")
        transaction_type = self.request.query_params.get("transaction_type")
        if master:
            queryset = queryset.filter(master_id=master)
        if transaction_type:
            queryset = queryset.filter(transaction_type=transaction_type)
        return queryset.order_by("-created_at")

    def perform_create(self, serializer):
        # A manual transaction must actually move the wallet balance, otherwise the
        # ledger and MasterWallet balances diverge.
        data = serializer.validated_data
        serializer.instance = post_wallet_transaction(
            master=data["master"],
            transaction_type=data["transaction_type"],
            amount=data["amount"],
            payment_method=data["payment_method"],
            description=data.get("description", ""),
            order=data.get("order"),
        )


@extend_schema(
    tags=[DASHBOARD_FINANCE_TAG],
    summary="Wallet tranzaksiya detail",
    description="Bitta wallet transaction yozuvini ko'rish, yangilash yoki o'chirish uchun ishlatiladi.",
)
class DashboardWalletTransactionDetailAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.RetrieveUpdateDestroyAPIView):
    serializer_class = DashboardWalletTransactionSerializer
    queryset = WalletTransaction.objects.select_related("master", "order").all()

    def perform_update(self, serializer):
        # Financial fields of a posted transaction are immutable (editing them would
        # desync the balance); only `description` may be corrected.
        instance = serializer.instance
        for field in ("master", "transaction_type", "amount", "payment_method", "order"):
            if field in serializer.validated_data and serializer.validated_data[field] != getattr(instance, field):
                raise ValidationError({field: LEDGER_IMMUTABLE_MSG})
        serializer.save()

    def perform_destroy(self, instance):
        # Deleting a transaction reverses its balance effect (guards negative balance).
        reverse_wallet_transaction(instance)
        instance.delete()


@extend_schema(
    tags=[DASHBOARD_FINANCE_TAG],
    summary="Pul yechish so'rovlari",
    description="Moliya sahifasida ustalarning pul yechish so'rovlarini status/master filterlari bilan qaytaradi yoki yangi so'rov yaratadi.",
)
class DashboardWithdrawRequestListCreateAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.ListCreateAPIView):
    serializer_class = DashboardWithdrawRequestSerializer

    def get_queryset(self):
        queryset = WithdrawRequest.objects.select_related("master")
        status = self.request.query_params.get("status")
        master = self.request.query_params.get("master")
        if status:
            queryset = queryset.filter(status=status)
        if master:
            queryset = queryset.filter(master_id=master)
        return queryset.order_by("-created_at")

    @transaction.atomic
    def perform_create(self, serializer):
        master = serializer.validated_data["master"]
        amount = serializer.validated_data["amount"]
        wallet, _ = MasterWallet.objects.select_for_update().get_or_create(master=master)
        pending_withdraw = WithdrawRequest.objects.filter(
            master=master,
            status=WithdrawRequest.PENDING,
        ).aggregate(amount=Sum("amount"))["amount"] or Decimal("0.00")
        if wallet.balance_cash - pending_withdraw < amount:
            raise ValidationError({"amount": "Naqd balans yetarli emas"})
        serializer.save()


@extend_schema(
    tags=[DASHBOARD_FINANCE_TAG],
    summary="Pul yechish so'rovi detail",
    description="Bitta withdraw requestni ko'rish, statusini yangilash yoki o'chirish uchun ishlatiladi.",
)
class DashboardWithdrawRequestDetailAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.RetrieveUpdateDestroyAPIView):
    serializer_class = DashboardWithdrawRequestSerializer
    queryset = WithdrawRequest.objects.select_related("master").all()

    def perform_update(self, serializer):
        instance = serializer.instance
        new_status = serializer.validated_data.get("status", instance.status)
        note = serializer.validated_data.get("admin_note") or ""
        # Route status transitions through the shared wallet service so approving a
        # withdraw here ACTUALLY debits the master's cash balance + records the OUT
        # WalletTransaction — exactly like the Unfold admin and the cash-handover
        # accept endpoint. Without this the endpoint only flipped the status and the
        # balance never moved. The service's PENDING guard keeps it idempotent
        # (re-PATCHing an already-approved request won't debit twice).
        if instance.status == WithdrawRequest.PENDING and new_status == WithdrawRequest.APPROVED:
            accept_cash_handover(instance, note=note)
            serializer.instance.refresh_from_db()
        elif instance.status == WithdrawRequest.PENDING and new_status == WithdrawRequest.REJECTED:
            reject_cash_handover(instance, note=note)
            serializer.instance.refresh_from_db()
        else:
            # A processed (approved/rejected) request's amount is settled — editing it
            # would desync the already-moved balance, so only allow it while pending.
            if instance.status != WithdrawRequest.PENDING and "amount" in serializer.validated_data:
                if serializer.validated_data["amount"] != instance.amount:
                    raise ValidationError({"amount": "Yakunlangan so'rov summasini o'zgartirib bo'lmaydi"})
            serializer.save()

    def perform_destroy(self, instance):
        # An approved request already debited the balance + wrote an OUT transaction;
        # deleting it would strand that money. Only a pending request may be deleted.
        if instance.status != WithdrawRequest.PENDING:
            raise ValidationError({"status": "Faqat 'pending' so'rovni o'chirish mumkin"})
        instance.delete()


@extend_schema(
    tags=[DASHBOARD_EXPENSES_TAG],
    summary="Kompaniya xarajatlari",
    description="Xarajatlar sahifasidagi kompaniya xarajatlari ro'yxatini date filterlar bilan qaytaradi yoki yangi xarajat yaratadi.",
)
class DashboardCompanyExpenseListCreateAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.ListCreateAPIView):
    serializer_class = DashboardCompanyExpenseSerializer

    def get_queryset(self):
        queryset = DashboardCompanyExpense.objects.all()
        date_from = self.request.query_params.get("date_from")
        date_to = self.request.query_params.get("date_to")
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)
        return queryset.order_by("-date", "-created_at")


@extend_schema(
    tags=[DASHBOARD_EXPENSES_TAG],
    summary="Kompaniya xarajati detail",
    description="Bitta kompaniya xarajatini ko'rish, tahrirlash yoki o'chirish uchun ishlatiladi.",
)
class DashboardCompanyExpenseDetailAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.RetrieveUpdateDestroyAPIView):
    serializer_class = DashboardCompanyExpenseSerializer
    queryset = DashboardCompanyExpense.objects.all()


@extend_schema(
    tags=[DASHBOARD_EXPENSES_TAG],
    summary="Ombor xarajatlari",
    description="Xarajatlar sahifasidagi ombor xarajatlarini product/date filterlar bilan qaytaradi yoki yangi ombor xarajati yaratadi.",
)
class DashboardWarehouseExpenseListCreateAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.ListCreateAPIView):
    serializer_class = DashboardWarehouseExpenseSerializer

    def get_queryset(self):
        queryset = DashboardWarehouseExpense.objects.select_related("product")
        product = self.request.query_params.get("product")
        date_from = self.request.query_params.get("date_from")
        date_to = self.request.query_params.get("date_to")
        if product:
            queryset = queryset.filter(product_id=product)
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)
        return queryset.order_by("-date", "-created_at")


@extend_schema(
    tags=[DASHBOARD_EXPENSES_TAG],
    summary="Ombor xarajati detail",
    description="Bitta ombor xarajatini ko'rish, tahrirlash yoki o'chirish uchun ishlatiladi.",
)
class DashboardWarehouseExpenseDetailAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.RetrieveUpdateDestroyAPIView):
    serializer_class = DashboardWarehouseExpenseSerializer
    queryset = DashboardWarehouseExpense.objects.select_related("product").all()


def _finance_int_param(request, name, default):
    """Parse an int query param, raising a 400 (not 500) on a bad value."""
    value = request.query_params.get(name)
    if value in (None, ""):
        return default
    try:
        return int(value)
    except (TypeError, ValueError) as exc:
        raise ValidationError({name: f"{name} butun son bo'lishi kerak"}) from exc


def _finance_month(request):
    """Return (year, month|None) from query params. Default: current month; `period=year` -> month=None."""
    today = default_target_date()
    year = _finance_int_param(request, "year", today.year)
    if request.query_params.get("period") == "year":
        return year, None
    month = _finance_int_param(request, "month", today.month)
    if not 1 <= month <= 12:
        raise ValidationError({"month": "month 1-12 oralig'ida bo'lishi kerak"})
    return year, month


def _finance_range(request):
    """Return (date_from, date_to, limit) for donut / top-clients.

    Default oraliq — tanlangan (yoki joriy) oy; `date_from`/`date_to` berilsa ustun turadi.
    """
    serializer = DateRangeQuerySerializer(data=request.query_params)
    serializer.is_valid(raise_exception=True)
    year, month = _finance_month(request)
    if month:
        month_start = date(year, month, 1)
        month_end = date(year, 12, 31) if month == 12 else date(year, month + 1, 1) - timedelta(days=1)
    else:
        month_start, month_end = date(year, 1, 1), date(year, 12, 31)
    date_from = serializer.validated_data.get("date_from") or month_start
    date_to = serializer.validated_data.get("date_to") or month_end
    return date_from, date_to, serializer.validated_data["limit"]


def _serialize_top_clients(clients, request):
    results = []
    for rank, client in enumerate(clients, start=1):
        avatar = None
        if client.avatar:
            try:
                avatar = request.build_absolute_uri(client.avatar.url)
            except ValueError:
                avatar = None
        results.append(
            {
                "rank": rank,
                "id": str(client.id),
                "full_name": str(client),
                "phone": client.phone,
                "avatar": avatar,
                "orders_count": client.completed_count,
                "total_spent": money(client.completed_spent),
                "total_spent_formatted": format_uzs(client.completed_spent),
            }
        )
    return results


@extend_schema(
    tags=[DASHBOARD_FINANCE_TAG],
    summary="Moliya summary",
    description=(
        "Moliya sahifasidagi Oylik daromad / Oylik xarajat / Sof foyda cardlari. Har biri qiymat, formatlangan "
        "matn va o'tgan davrga nisbatan o'zgarish foizini (change_percent/direction) qaytaradi. "
        "Default `period=month` (joriy oy). `period=year` — yillik. `year` va `month` bilan davr tanlanadi."
    ),
    parameters=[
        OpenApiParameter("period", str, OpenApiParameter.QUERY, required=False, description="month (default) yoki year."),
        OpenApiParameter("year", int, OpenApiParameter.QUERY, required=False),
        OpenApiParameter("month", int, OpenApiParameter.QUERY, required=False, description="1-12 (period=month uchun)."),
    ],
)
class DashboardFinanceSummaryAPIView(DashboardPermissionMixin, generics.GenericAPIView):
    serializer_class = EmptySerializer

    def get(self, request):
        year, month = _finance_month(request)
        return success_response(get_finance_summary(year, month))


@extend_schema(
    tags=[DASHBOARD_FINANCE_TAG],
    summary="Daromad xizmatlar kesimida (donut)",
    description=(
        "'Oylik daromada xizmatlar kesimida' donut charti uchun yakunlangan buyurtmalar daromadini "
        "(Sum total_amount) xizmatlar bo'yicha taqsimlab qaytaradi. Default oraliq — joriy oy."
    ),
    parameters=[
        OpenApiParameter("year", int, OpenApiParameter.QUERY, required=False),
        OpenApiParameter("month", int, OpenApiParameter.QUERY, required=False),
        OpenApiParameter("date_from", str, OpenApiParameter.QUERY, required=False),
        OpenApiParameter("date_to", str, OpenApiParameter.QUERY, required=False),
        OpenApiParameter("limit", int, OpenApiParameter.QUERY, required=False),
    ],
)
class DashboardIncomeByServiceAPIView(DashboardPermissionMixin, generics.GenericAPIView):
    serializer_class = DateRangeQuerySerializer

    def get(self, request):
        date_from, date_to, limit = _finance_range(request)
        return success_response(get_income_by_service(date_from, date_to, limit))


@extend_schema(
    tags=[DASHBOARD_FINANCE_TAG],
    summary="Eng faol mijozlar",
    description=(
        "'Eng faol mijozlar' jadvali uchun yakunlangan buyurtmalar soni va sarflagan summasi bo'yicha "
        "top mijozlarni qaytaradi. Default oraliq — joriy oy; `date_from`/`date_to` bilan o'zgartirish mumkin."
    ),
    parameters=[
        OpenApiParameter("year", int, OpenApiParameter.QUERY, required=False),
        OpenApiParameter("month", int, OpenApiParameter.QUERY, required=False),
        OpenApiParameter("date_from", str, OpenApiParameter.QUERY, required=False),
        OpenApiParameter("date_to", str, OpenApiParameter.QUERY, required=False),
        OpenApiParameter("limit", int, OpenApiParameter.QUERY, required=False, description="Nechta mijoz (default 10)."),
    ],
)
class DashboardTopClientsAPIView(DashboardPermissionMixin, generics.GenericAPIView):
    serializer_class = DateRangeQuerySerializer

    def get(self, request):
        date_from, date_to, limit = _finance_range(request)
        clients = get_top_clients(date_from, date_to, limit)
        results = _serialize_top_clients(clients, request)
        return success_response({"count": len(results), "results": results})


@extend_schema(
    tags=[DASHBOARD_FINANCE_TAG],
    summary="Moliya hisobot jadvali",
    description=(
        "Moliya sahifasi uchun bitta so'rovda summary cardlar, oyma-oy income/expense chart, xizmatlar "
        "kesimidagi daromad (donut) va eng faol mijozlar ma'lumotlarini qaytaradi."
    ),
)
class DashboardFinanceReportAPIView(DashboardPermissionMixin, generics.GenericAPIView):
    serializer_class = EmptySerializer

    def get(self, request):
        year, month = _finance_month(request)
        date_from, date_to, limit = _finance_range(request)
        return success_response(
            {
                "summary": get_finance_summary(year, month),
                "chart": get_income_expense(year),
                "income_by_service": get_income_by_service(date_from, date_to, limit),
                "top_clients": _serialize_top_clients(get_top_clients(date_from, date_to, limit), request),
            }
        )


@extend_schema(
    tags=[DASHBOARD_FINANCE_TAG],
    summary="Moliya hisobotni Excel'ga eksport qilish",
    description=(
        "'Export Excel' tugmasi uchun. Summary, oyma-oy daromad/xarajat, xizmatlar kesimidagi daromad va "
        "eng faol mijozlarni `.xlsx` fayl ko'rinishida yuklab beradi."
    ),
    parameters=[
        OpenApiParameter("year", int, OpenApiParameter.QUERY, required=False),
        OpenApiParameter("month", int, OpenApiParameter.QUERY, required=False),
        OpenApiParameter("period", str, OpenApiParameter.QUERY, required=False),
    ],
)
class DashboardFinanceExportAPIView(DashboardPermissionMixin, generics.GenericAPIView):
    serializer_class = EmptySerializer

    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font

        year, month = _finance_month(request)
        date_from, date_to, limit = _finance_range(request)
        summary = get_finance_summary(year, month)
        chart = get_income_expense(year)
        by_service = get_income_by_service(date_from, date_to, limit)
        top_clients = get_top_clients(date_from, date_to, limit)
        bold = Font(bold=True)

        workbook = Workbook()

        overview = workbook.active
        overview.title = "Umumiy"
        overview.append(["Davr", summary["label"]])
        overview.append(["Ko'rsatkich", "Summa (so'm)", "O'zgarish %", "Yo'nalish"])
        overview["A2"].font = bold
        overview["B2"].font = bold
        overview["C2"].font = bold
        overview["D2"].font = bold
        for key, title in (("income", "Daromad"), ("expense", "Xarajat"), ("profit", "Sof foyda")):
            card = summary[key]
            overview.append([title, card["value"], card["change_percent"], card["change_direction"]])
        overview.append(["Yakunlangan buyurtmalar", summary["orders_count"]])
        overview.append(["Tasdiqlanmagan naqd so'rovlar", summary["withdraw_pending_count"]])

        monthly = workbook.create_sheet("Oyma-oy")
        monthly.append(["Oy", "Daromad (mln)", "Xarajat (mln)"])
        for cell in ("A1", "B1", "C1"):
            monthly[cell].font = bold
        for item in chart["items"]:
            monthly.append([item["label"], item["income"], item["expense"]])

        services_sheet = workbook.create_sheet("Xizmatlar kesimida")
        services_sheet.append(["Xizmat", "Daromad (so'm)", "Ulush %"])
        for cell in ("A1", "B1", "C1"):
            services_sheet[cell].font = bold
        for item in by_service["items"]:
            services_sheet.append([item["service_name"], item["income"], item["percent"]])

        clients_sheet = workbook.create_sheet("Eng faol mijozlar")
        clients_sheet.append(["#", "Mijoz", "Telefon", "Buyurtmalar", "Sarflagan (so'm)"])
        for cell in ("A1", "B1", "C1", "D1", "E1"):
            clients_sheet[cell].font = bold
        for rank, client in enumerate(top_clients, start=1):
            clients_sheet.append(
                [rank, str(client), client.phone, client.completed_count, money(client.completed_spent)]
            )

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        suffix = f"{year}-{month:02d}" if month else str(year)
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="moliya-hisobot-{suffix}.xlsx"'
        return response


@extend_schema(
    tags=[DASHBOARD_SETTINGS_TAG],
    summary="Integratsiya sozlamalari",
    description="Sozlamalar sahifasidagi integratsiya kalitlari va konfiguratsiyalarini qaytaradi yoki yangi integration setting yaratadi.",
)
class DashboardIntegrationSettingListCreateAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.ListCreateAPIView):
    serializer_class = DashboardIntegrationSettingSerializer

    def get_queryset(self):
        queryset = DashboardIntegrationSetting.objects.all()
        is_active = bool_param(self.request.query_params.get("is_active"))
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active)
        return queryset.order_by("key")


@extend_schema(
    tags=[DASHBOARD_SETTINGS_TAG],
    summary="Integratsiya sozlamasi detail",
    description="Bitta integration settingni ko'rish, qiymatini yangilash yoki o'chirish uchun ishlatiladi.",
)
class DashboardIntegrationSettingDetailAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.RetrieveUpdateDestroyAPIView):
    serializer_class = DashboardIntegrationSettingSerializer
    queryset = DashboardIntegrationSetting.objects.all()


@extend_schema(
    tags=[DASHBOARD_SETTINGS_TAG],
    summary="Backup ro'yxati / yaratish",
    description=(
        "GET: mavjud DB backuplari ro'yxati (eng yangisi birinchi). "
        "POST: hozir to'liq DB backup (.sql) yaratadi — DB yiqilsa shu fayldan tiklash mumkin."
    ),
)
class DashboardBackupListCreateAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.ListCreateAPIView):
    serializer_class = DashboardBackupSerializer

    def get_queryset(self):
        return DashboardBackup.objects.all()

    def create(self, request, *args, **kwargs):
        try:
            backup = create_backup(
                source=DashboardBackup.MANUAL, created_by=request.user, note=request.data.get("note", "")
            )
        except RuntimeError as exc:
            return Response({"success": False, "message": str(exc)}, status=500)
        data = DashboardBackupSerializer(backup, context={"request": request}).data
        return success_response(data, status=201, message="Backup yaratildi")


@extend_schema(
    tags=[DASHBOARD_SETTINGS_TAG],
    summary="Backup detail / o'chirish",
    description="Bitta backupni ko'rish yoki o'chirish (fayl ham diskdan o'chadi).",
)
class DashboardBackupDetailAPIView(DashboardPermissionMixin, EnvelopeMixin, generics.RetrieveDestroyAPIView):
    serializer_class = DashboardBackupSerializer
    queryset = DashboardBackup.objects.all()

    def perform_destroy(self, instance):
        delete_backup(instance)


@extend_schema(
    tags=[DASHBOARD_SETTINGS_TAG],
    summary="Backup faylni yuklab olish",
    description="Backup `.sql` faylini yuklab oladi (faqat admin). Fayldan `psql`/`sqlite3` orqali DB tiklanadi.",
)
class DashboardBackupDownloadAPIView(DashboardPermissionMixin, generics.GenericAPIView):
    serializer_class = EmptySerializer

    def get(self, request, pk):
        backup = get_object_or_404(DashboardBackup, pk=pk)
        if not backup.exists:
            raise Http404("Backup fayli topilmadi")
        return FileResponse(
            open(backup.path, "rb"),
            as_attachment=True,
            filename=backup.filename,
            content_type="application/sql",
        )


@extend_schema(
    tags=[DASHBOARD_SETTINGS_TAG],
    summary="Avtomatik backup sozlamasi",
    description="GET: joriy sozlama. PATCH: `enabled`, `day_of_week` (0-6), `hour`, `keep` o'zgartirish. Avto backup haftalik.",
)
class DashboardBackupSettingsAPIView(DashboardPermissionMixin, generics.GenericAPIView):
    serializer_class = DashboardBackupSettingsSerializer

    def get(self, request):
        return success_response(get_auto_config())

    def patch(self, request):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        config = set_auto_config(**serializer.validated_data)
        return success_response(config, message="Avtomatik backup sozlamasi yangilandi")

    post = patch
